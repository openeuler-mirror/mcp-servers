from __future__ import annotations

import logging
import os
import copy
import re
import sys
import time

import git
import yaml
from dataclasses import dataclass

from .backporting import run_backport_from_config
from .commit_message_template import (
    DEFAULT_COMMIT_MESSAGE_TEMPLATE,
    DEFAULT_LINUX_REPO_PATH,
    build_commit_message_preview,
    normalize_commit_message_source,
)
from .locales import i18n
from .backport_sort import resolve_sorted_backport_items
from tabulate import tabulate

logger = logging.getLogger(__name__)

import readline


def _resolve_commit_message_source(args, item_config: dict, base_config: dict) -> str:
    return normalize_commit_message_source(
        str(getattr(args, "commit_message_source", "") or "").strip()
        or str(item_config.get("commit_message_source") or "").strip()
        or str(base_config.get("commit_message_source") or "").strip()
        or "auto"
    )


def _resolve_backport_engine(args, item_config: dict, base_config: dict) -> str:
    engine = (
        str(item_config.get("backport_engine") or "").strip().lower()
        or str(getattr(args, "backport_engine", "") or "").strip().lower()
        or str(base_config.get("backport_engine") or "").strip().lower()
        or "portgpt"
    )
    if engine not in {"portgpt", "mystique"}:
        raise ValueError(
            f"不支持的 backport_engine: {engine!r}，请选择 portgpt 或 mystique"
        )
    return engine


@dataclass
class BackportBatchContext:
    config: dict
    is_report_config: bool
    report_output_path: str
    base_config: dict
    base_project_dir: str
    base_target_path: str
    sorted_items: list
    sort_errors: list


def generate_backport_batch_config_from_excel(
    *,
    excel_path: str,
    output_path: str,
    template_config_path: str | None = None,
    sheet_name: str | None = None,
):
    """根据 Excel 生成 backport-batch 原始配置文件。"""
    if not excel_path:
        raise ValueError("请提供 Excel 文件路径")
    if not os.path.exists(excel_path):
        raise ValueError(f"Excel 文件不存在: {excel_path}")
    if not output_path:
        raise ValueError("请通过 -o/--output 指定输出配置文件路径")

    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise ValueError(
            "缺少 openpyxl 依赖，无法读取 Excel。请先安装: pip install openpyxl"
        ) from e

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"指定 sheet 不存在: {sheet_name}，可选: {workbook.sheetnames}"
            )
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook[workbook.sheetnames[0]]

    rows = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        raise ValueError("Excel 内容为空，无法生成配置")

    title_idx, hash_idx = _resolve_excel_commit_columns(header_row)
    commits = []
    for row in rows:
        if not row:
            continue
        commit_title = _normalize_excel_cell_value(row, title_idx)
        commit_hash = _normalize_excel_cell_value(row, hash_idx)
        if not commit_title and not commit_hash:
            continue
        commit_item = {}
        if commit_hash:
            commit_item["commit"] = commit_hash
        if commit_title:
            commit_item["commit_title"] = commit_title
        commits.append(commit_item)

    if not commits:
        raise ValueError("Excel 中未解析到有效提交记录（commit title / commit hash）")

    base_config = {}
    if template_config_path:
        if not os.path.exists(template_config_path):
            raise ValueError(f"模板配置文件不存在: {template_config_path}")
        with open(template_config_path, "r", encoding="utf-8") as file:
            loaded = yaml.safe_load(file) or {}
        if not isinstance(loaded, dict):
            raise ValueError("模板配置必须是对象/字典结构")
        base_config = {k: v for k, v in loaded.items() if k != "commits"}

    base_config["commits"] = commits
    with open(output_path, "w", encoding="utf-8") as file:
        yaml.safe_dump(base_config, file, allow_unicode=True, sort_keys=False)

    return {
        "status": "success",
        "action": "generate-backport-batch-config",
        "excel_path": excel_path,
        "sheet_name": worksheet.title,
        "output_path": output_path,
        "commit_count": len(commits),
        "template_config_path": template_config_path or "",
    }


def _normalize_excel_header(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    return "".join(ch for ch in text if ch.isalnum())


def _resolve_excel_commit_columns(header_row):
    if not header_row:
        raise ValueError("Excel 首行为空，无法识别列名")
    normalized_headers = [_normalize_excel_header(value) for value in header_row]

    title_aliases = {"committitle", "title", "subject", "patchtitle"}
    hash_aliases = {"commithash", "commit", "hash", "sha", "commitid"}

    title_idx = None
    hash_idx = None
    for idx, normalized in enumerate(normalized_headers):
        if title_idx is None and normalized in title_aliases:
            title_idx = idx
        if hash_idx is None and normalized in hash_aliases:
            hash_idx = idx

    if title_idx is None or hash_idx is None:
        raise ValueError(
            "Excel 缺少必需列，请包含 commit title 与 commit hash。"
            f"检测到列: {list(header_row)}"
        )
    return title_idx, hash_idx


def _normalize_excel_cell_value(row, idx) -> str:
    if idx is None or idx >= len(row):
        return ""
    value = row[idx]
    if value is None:
        return ""
    return str(value).strip()


def handle_backport_batch(args):
    """处理批量补丁回移植逻辑（从 cli 入口拆分到 utils 模块）"""
    if getattr(args, "preview_commit_message", False):
        return _handle_preview_commit_message(args)
    if getattr(args, "apply", None):
        return _handle_direct_apply_backported_patch(args)
    logger.info("[backport-batch] 开始处理: config=%s", args.backport_config)
    execute_requested = bool(getattr(args, "execute", False))
    config_path = str(getattr(args, "backport_config", "") or "")
    if execute_requested and not config_path.endswith(".report.yml"):
        raise ValueError(
            "检测到 --execute/-e 但当前配置不是 .report.yml。"
            "请先使用 raw 配置生成 report，再对 .report.yml 执行回移植。"
        )
    context = _prepare_backport_batch_context(args)
    if context is None:
        logger.info("[backport-batch] 用户在交互模式中选择退出，停止执行")
        return []
    is_report_config = context.is_report_config
    if execute_requested and not is_report_config:
        raise ValueError(
            "检测到 --execute/-e 但当前配置不是 .report.yml。"
            "请先使用 raw 配置生成 report，再对 .report.yml 执行回移植。"
        )
    if is_report_config and not execute_requested and not getattr(args, "stop_at_first_conflict", False):
        logger.warning(
            "[backport-batch] 检测到 .report.yml 但未显式指定 --execute/-e，"
            "沿用兼容模式继续执行。建议显式加 -e。"
        )
    base_config = context.base_config
    base_project_dir = context.base_project_dir
    base_target_path = context.base_target_path
    sorted_items = context.sorted_items
    sort_errors = context.sort_errors
    report_output_path = context.report_output_path or args.backport_config

    default_target_branch = args.branch
    results, report_items = _execute_backport_batch_items(
        sorted_items=sorted_items,
        sort_errors=sort_errors,
        is_report_config=is_report_config,
        base_config=base_config,
        base_project_dir=base_project_dir,
        base_target_path=base_target_path,
        default_target_branch=default_target_branch,
        args=args,
    )

    report = _build_backport_batch_report(
        base_config=base_config,
        base_project_dir=base_project_dir,
        base_target_path=base_target_path,
        default_target_branch=default_target_branch,
        llm_provider=args.llm_provider,
        llm_base_url=getattr(args, 'llm_base_url', None),
        llm_model_name=getattr(args, 'llm_model_name', None),
        backport_engine=_resolve_backport_engine(args, {}, base_config),
        report_items=report_items,
    )
    _write_backport_batch_report(report_output_path, is_report_config, report)
    return results


def _build_commit_message_report_fields(
    *,
    patch_path: str,
    openeuler_commit_id: str,
    item_config: dict,
    base_config: dict,
    args,
) -> dict:
    if not patch_path or not os.path.exists(patch_path):
        return {
            "commit_message_preview": "",
            "commit_message_context": {},
            "source_detection": {
                "source": "openEuler",
                "commit_id": str(openeuler_commit_id or ""),
                "method": "missing_patch",
                "warning": "缺少原始 patch，无法生成 commit message 预览。",
            },
            "commit_message_warnings": ["缺少原始 patch，无法生成 commit message 预览。"],
        }
    try:
        return build_commit_message_preview(
            patch_path=patch_path,
            openeuler_commit_id=openeuler_commit_id,
            template=(
                str(getattr(args, "commit_message_template", "") or "").strip()
                or str(item_config.get("commit_message_template") or "").strip()
                or str(base_config.get("commit_message_template") or "").strip()
                or DEFAULT_COMMIT_MESSAGE_TEMPLATE
            ),
            linux_repo_path=(
                str(getattr(args, "linux_repo_path", "") or "").strip()
                or str(item_config.get("linux_repo_path") or "").strip()
                or str(base_config.get("linux_repo_path") or "").strip()
                or DEFAULT_LINUX_REPO_PATH
            ),
            commit_message_source=_resolve_commit_message_source(args, item_config, base_config),
        )
    except Exception as exc:
        logger.warning("[backport-batch] commit message preview failed: %s", exc)
        return {
            "commit_message_preview": "",
            "commit_message_context": {},
            "source_detection": {
                "source": "openEuler",
                "commit_id": str(openeuler_commit_id or ""),
                "method": "preview_failed",
                "warning": str(exc),
            },
            "commit_message_warnings": [str(exc)],
        }


def _resolve_commit_message_fields(
    *,
    patch_path: str,
    openeuler_commit_id: str,
    item_config: dict,
    base_config: dict,
    args,
) -> dict:
    existing_preview = str(item_config.get("commit_message_preview") or "").strip()
    if existing_preview:
        return {
            "commit_message": existing_preview,
            "commit_message_preview": existing_preview,
            "commit_message_context": item_config.get("commit_message_context") or {},
            "source_detection": item_config.get("source_detection") or {},
            "commit_message_warnings": item_config.get("commit_message_warnings") or [],
        }
    return _build_commit_message_report_fields(
        patch_path=patch_path,
        openeuler_commit_id=openeuler_commit_id,
        item_config=item_config,
        base_config=base_config,
        args=args,
    )


def _looks_like_patch_path(value: str) -> bool:
    if not isinstance(value, str):
        return False
    return value.strip().endswith(".patch") or os.path.sep in value or value.startswith(".")


def _resolve_item_config(item, is_report_config):
    return item if is_report_config else item.get("item_config", {})


def _select_apply_item(sorted_items, is_report_config, apply_value: str):
    query = (apply_value or "").strip()
    if not query:
        raise ValueError("--apply 不能为空")

    normalized_query = os.path.abspath(os.path.expanduser(query)) if _looks_like_patch_path(query) else query
    by_path = _looks_like_patch_path(query)
    matches = []
    for item in sorted_items:
        item_config = _resolve_item_config(item, is_report_config)
        if by_path:
            candidate_paths = [
                item_config.get("backported_patch_path"),
                item_config.get("patch_path"),
                item_config.get("original_patch_path"),
            ]
            normalized_candidates = []
            for candidate_path in candidate_paths:
                if not candidate_path:
                    continue
                normalized_candidates.append(
                    os.path.abspath(os.path.expanduser(str(candidate_path)))
                )
            if normalized_query in normalized_candidates:
                matches.append(item)
        else:
            commit_candidates = [
                str(item.get("commit") or "").strip(),
                str(item.get("input_commit") or "").strip(),
            ]
            if any(candidate.lower().startswith(query.lower()) for candidate in commit_candidates if candidate):
                matches.append(item)

    if not matches:
        mode = "patch路径" if by_path else "commit id"
        raise ValueError(f"在配置中未找到匹配的 {mode}: {query}")
    if len(matches) > 1:
        raise ValueError(f"--apply 匹配到多个提交，请提供更精确值: {query}")
    return matches[0]


def _handle_preview_commit_message(args):
    logger.info("[backport-batch] 进入 commit message 预览模式: apply=%s", args.apply)
    if not getattr(args, "apply", None):
        raise ValueError("--preview-commit-message 需要同时提供 --apply")
    context = _prepare_backport_batch_context(args)
    if context is None:
        return {
            "status": "cancelled",
            "action": "backport-batch-preview-commit-message",
            "message": "用户在交互模式中退出",
        }

    item = _select_apply_item(
        context.sorted_items,
        context.is_report_config,
        args.apply,
    )
    item_config = _resolve_item_config(item, context.is_report_config)
    commit_id = str(item.get("commit") or item.get("input_commit") or "").strip()
    apply_value = str(args.apply).strip()
    if _looks_like_patch_path(apply_value):
        patch_path = os.path.abspath(os.path.expanduser(apply_value))
    else:
        patch_path = (
            item_config.get("backported_patch_path")
            or item_config.get("patch_path")
            or item_config.get("original_patch_path")
            or ""
        )
    original_patch_path = _infer_original_patch_path(
        item_config=item_config,
    )
    commit_message_source = original_patch_path or patch_path
    if not commit_message_source:
        raise ValueError(f"未找到可预览的补丁路径: commit={commit_id}")
    preview = _resolve_commit_message_fields(
        patch_path=commit_message_source,
        openeuler_commit_id=commit_id,
        item_config=item_config,
        base_config=context.base_config,
        args=args,
    )
    return {
        "action": "backport-batch-preview-commit-message",
        "status": "success",
        "apply": args.apply,
        "commit": commit_id,
        "patch_path": patch_path,
        "original_patch_path": original_patch_path,
        "commit_message_source": commit_message_source,
        **preview,
    }


def _infer_original_patch_path(item_config: dict):
    for key in ("original_patch_path", "patch_path", "backported_patch_path"):
        value = str(item_config.get(key) or "").strip()
        if not value:
            continue
        path = os.path.abspath(os.path.expanduser(value))
        if os.path.exists(path):
            return path
    return ""


def _parse_patch_author(patch_path: str) -> tuple[str, str]:
    try:
        with open(patch_path, "r", encoding="utf-8", errors="ignore") as handle:
            for _ in range(80):
                line = handle.readline()
                if not line:
                    break
                if not line.strip():
                    break
                match = re.match(r"^(?:From|Author):\s*(.*?)\s*<([^<>]+)>\s*$", line.strip())
                if match:
                    return match.group(1).strip(), match.group(2).strip()
    except OSError:
        return "", ""
    return "", ""


def _is_pending_backport_item(item) -> bool:
    if not isinstance(item, dict):
        return False
    return str(item.get("status") or "").strip().lower() == "pending"


def _copy_existing_report_item(item):
    if isinstance(item, dict):
        return copy.deepcopy(item)
    return item


def _apply_patch_file_to_target_repo(
    *,
    target_path: str,
    target_branch: str,
    patch_path: str,
    commit_message: str,
    signer_name: str = "",
    signer_email: str = "",
):
    if not patch_path or not os.path.exists(patch_path):
        return {"status": "failed", "error": f"补丁文件不存在: {patch_path}"}

    target_repo = git.Repo(target_path)
    _ensure_clean_and_checkout(target_repo, target_branch)
    original_head = target_repo.head.commit.hexsha

    try:
        already_applied, reverse_check_error = _is_patch_applied_in_target(
            target_path, target_branch, patch_path
        )
        if already_applied:
            return {"status": "skipped", "error": "目标分支已包含该补丁"}

        target_repo.git.apply(patch_path)
        target_repo.git.add("--all")
        if not target_repo.is_dirty(index=True, working_tree=True, untracked_files=False):
            return {"status": "skipped", "error": "补丁未产生变更（可能已等效存在）"}

        commit_args = ["-m", commit_message, "-s"]
        use_signer_identity = bool(signer_name and signer_email)
        author_name, author_email = _parse_patch_author(patch_path)
        git_env = {}
        if author_name and author_email:
            git_env["GIT_AUTHOR_NAME"] = author_name
            git_env["GIT_AUTHOR_EMAIL"] = author_email
        if use_signer_identity:
            git_env["GIT_COMMITTER_NAME"] = signer_name
            git_env["GIT_COMMITTER_EMAIL"] = signer_email
        if git_env:
            with target_repo.git.custom_environment(**git_env):
                target_repo.git.commit(*commit_args)
        else:
            target_repo.git.commit(*commit_args)
        return {
            "status": "success",
            "commit": target_repo.head.commit.hexsha,
        }
    except Exception as e:
        _reset_hard_and_clean(target_repo, original_head)
        if reverse_check_error:
            return {"status": "failed", "error": f"{e}; reverse-check={reverse_check_error}"}
        return {"status": "failed", "error": str(e)}


def _handle_direct_apply_backported_patch(args):
    logger.info("[backport-batch] 进入直接 apply 模式: apply=%s", args.apply)
    context = _prepare_backport_batch_context(args)
    if context is None:
        return {
            "status": "cancelled",
            "action": "backport-batch-apply",
            "message": "用户在交互模式中退出",
        }

    item = _select_apply_item(
        context.sorted_items,
        context.is_report_config,
        args.apply,
    )
    item_config = _resolve_item_config(item, context.is_report_config)
    commit_id = item.get("commit") or item.get("input_commit") or ""
    target_branch = _resolve_target_branch(item_config, context.base_config, args.branch)
    if not target_branch:
        raise ValueError("未找到目标分支，请在配置中设置 target_release/target_branch 或传入 --branch")

    apply_value = str(args.apply).strip()
    apply_by_path = _looks_like_patch_path(apply_value)
    if apply_by_path:
        patch_path = os.path.abspath(os.path.expanduser(apply_value))
    else:
        patch_path = (
            item_config.get("backported_patch_path")
            or item_config.get("patch_path")
            or item_config.get("original_patch_path")
            or ""
        )

    if not patch_path:
        raise ValueError(f"未找到可应用的补丁路径: commit={commit_id}")
    if not os.path.exists(patch_path):
        raise ValueError(f"补丁文件不存在: {patch_path}")

    original_patch_path = _infer_original_patch_path(
        item_config=item_config,
    )
    commit_message_source = original_patch_path or patch_path
    commit_message_preview = _resolve_commit_message_fields(
        patch_path=commit_message_source,
        openeuler_commit_id=str(commit_id or ""),
        item_config=item_config,
        base_config=context.base_config,
        args=args,
    )
    commit_message = commit_message_preview["commit_message"]
    apply_result = _apply_patch_file_to_target_repo(
        target_path=context.base_target_path,
        target_branch=target_branch,
        patch_path=patch_path,
        commit_message=commit_message,
        signer_name=str(getattr(args, "signer_name", "") or context.base_config.get("signer_name") or "").strip(),
        signer_email=str(getattr(args, "signer_email", "") or context.base_config.get("signer_email") or "").strip(),
    )
    if apply_result.get("status") == "success":
        _write_apply_status_to_report_config(
            config_path=args.backport_config,
            config=context.config,
            selected_item=item,
            selected_item_config=item_config,
            selected_commit_id=str(commit_id or ""),
            applied_patch_path=patch_path,
            target_branch=target_branch,
            applied_commit=apply_result.get("commit"),
        )
    return {
        "action": "backport-batch-apply",
        "status": apply_result.get("status"),
        "apply": args.apply,
        "commit": commit_id,
        "target_branch": target_branch,
        "patch_path": patch_path,
        "original_patch_path": original_patch_path,
        "commit_message_source": commit_message_source,
        "commit_message_preview": commit_message_preview.get("commit_message_preview", ""),
        "commit_message_context": commit_message_preview.get("commit_message_context", {}),
        "source_detection": commit_message_preview.get("source_detection", {}),
        "commit_message_warnings": commit_message_preview.get("commit_message_warnings", []),
        "applied_commit": apply_result.get("commit"),
        "error": apply_result.get("error"),
    }


def _write_apply_status_to_report_config(
    *,
    config_path: str,
    config: dict,
    selected_item: dict,
    selected_item_config: dict,
    selected_commit_id: str,
    applied_patch_path: str,
    target_branch: str,
    applied_commit: str | None,
):
    if not isinstance(config, dict):
        return
    commit_items = config.get("commits")
    if not isinstance(commit_items, list):
        return

    selected_input_commit = str(selected_item.get("input_commit") or "").strip()
    selected_paths = set()
    for path_value in [
        applied_patch_path,
        selected_item_config.get("backported_patch_path"),
        selected_item_config.get("patch_path"),
        selected_item_config.get("original_patch_path"),
    ]:
        if not path_value:
            continue
        selected_paths.add(os.path.abspath(os.path.expanduser(str(path_value))))

    matched_index = None
    for idx, item in enumerate(commit_items):
        if not isinstance(item, dict):
            continue
        item_commit = str(item.get("commit") or "").strip()
        item_input_commit = str(item.get("input_commit") or "").strip()
        commit_matched = False
        if selected_commit_id and item_commit and item_commit.lower().startswith(selected_commit_id.lower()):
            commit_matched = True
        elif selected_commit_id and selected_commit_id.lower().startswith(item_commit.lower()) and item_commit:
            commit_matched = True
        elif selected_input_commit and item_input_commit and item_input_commit.lower() == selected_input_commit.lower():
            commit_matched = True

        path_matched = False
        for candidate in [
            item.get("backported_patch_path"),
            item.get("patch_path"),
            item.get("original_patch_path"),
        ]:
            if not candidate:
                continue
            normalized = os.path.abspath(os.path.expanduser(str(candidate)))
            if normalized in selected_paths:
                path_matched = True
                break

        if commit_matched or path_matched:
            matched_index = idx
            break

    if matched_index is None:
        return

    matched_item = commit_items[matched_index]
    matched_item["merged_in_target"] = True
    matched_item["merged_check_error"] = None
    matched_item["has_conflict"] = False
    matched_item["conflict_check_method"] = "apply"
    matched_item["conflict_check_error"] = None
    matched_item["status"] = "success"
    matched_item["error"] = None
    matched_item["patch_path"] = applied_patch_path
    matched_item["target_branch"] = target_branch
    if applied_commit:
        matched_item["applied_commit"] = applied_commit

    with open(config_path, "w", encoding="utf-8") as f:
        yaml.safe_dump(config, f, allow_unicode=True, sort_keys=False)


def _load_backport_batch_config(config_path: str):
    logger.info("[backport-batch] 读取配置: path=%s", config_path)
    if not config_path:
        raise ValueError("backport-batch 需要提供 --backport-config")
    if not os.path.exists(config_path):
        raise ValueError(f"配置文件不存在: {config_path}")
    with open(config_path, "r") as file:
        config = yaml.safe_load(file)
    if not isinstance(config, dict):
        raise ValueError("配置文件内容必须是对象/字典结构")
    commit_items = _normalize_backport_batch_commits(config.get("commits", []))
    if not isinstance(commit_items, list) or not commit_items:
        raise ValueError("配置文件必须包含非空的 commits 列表/字典")
    logger.info(
        "[backport-batch] 配置读取完成: commits=%d, keys=%s",
        len(commit_items),
        list(config.keys()),
    )
    return config, commit_items


def _normalize_backport_batch_commits(commit_items):
    if not isinstance(commit_items, dict):
        return commit_items

    normalized_items = []
    for commit_key, commit_value in commit_items.items():
        if isinstance(commit_value, dict):
            normalized = dict(commit_value)
            normalized.setdefault("commit", commit_key)
        else:
            normalized = {"commit": commit_key, "commit_title": commit_value}
        normalized_items.append(normalized)
    return normalized_items

def _is_report_config(config_path: str, commit_items):
    if isinstance(config_path, str) and config_path.endswith(".report.yml"):
        return True
    # 非 .report.yml 一律视为原始配置：
    # 即使条目中包含 report 字段（例如复制自 report 或手工补充），也不进入 report 执行路径，
    # 只生成/刷新对应的 .report.yml，避免误触发补丁应用。
    if isinstance(commit_items, list):
        for item in commit_items:
            if not isinstance(item, dict):
                continue
            if (
                "has_conflict" in item
                or "conflict_check_method" in item
                or "merged_in_target" in item
            ):
                logger.info(
                    "[backport-batch] 检测到 report 字段但配置文件非 .report.yml，仍按 raw 处理: config=%s",
                    config_path,
                )
                break
    return False

def _extract_commit_item(item):
    commit_id = None
    commit_title = None
    item_config = {}
    if isinstance(item, str):
        commit_id = item
    elif isinstance(item, (list, tuple)) and len(item) >= 2:
        commit_id = item[0]
        commit_title = item[1]
    elif isinstance(item, dict):
        commit_id = (
            item.get("commit")
            or item.get("commit_id")
            or item.get("sha")
            or item.get("new_patch")
        )
        commit_title = item.get("commit_title") or item.get("title") or item.get("subject")
        item_config = item
    return commit_id, commit_title, item_config

def _parse_merged_in_target_value(value: str, *, allow_skipped: bool = False):
    if value is None:
        return None
    text = str(value).strip().lower()
    if text in {"1", "true", "t", "yes", "y"}:
        return True
    if text in {"0", "false", "f", "no", "n"}:
        return False
    if text in {"none", "null", "na", "n/a", "-"}:
        return None
    if allow_skipped and text == "skipped":
        return "skipped"
    return "invalid"


def _interactive_input(prompt_text: str) -> str:
    """在可用时启用 readline 行编辑和历史能力。"""
    if readline is not None:
        try:
            readline.parse_and_bind("set editing-mode emacs")
            readline.parse_and_bind("tab: complete")
        except Exception:
            pass
    text = input(prompt_text)
    if readline is not None:
        try:
            stripped = text.strip()
            if stripped:
                readline.add_history(stripped)
        except Exception:
            pass
    return text


def _parse_bool_text(value: str):
    if value is None:
        return "invalid"
    text = str(value).strip().lower()
    if text in {"1", "true", "t", "yes", "y"}:
        return True
    if text in {"0", "false", "f", "no", "n"}:
        return False
    return "invalid"


def _match_commit_item_with_query(item: dict, query: str):
    if not isinstance(item, dict):
        return False, "条目格式无效"
    raw_query = (query or "").strip()
    if not raw_query:
        return False, "搜索条件不能为空"

    key = ""
    value = raw_query
    if ":" in raw_query:
        key, value = raw_query.split(":", 1)
    elif "=" in raw_query:
        key, value = raw_query.split("=", 1)
    key = key.strip().lower()
    value = value.strip()

    if key in {"", "title", "tiltle", "commit_title"}:
        title = str(item.get("commit_title") or "")
        return title.lower().startswith(value.lower()), None

    if key in {"has_conflict", "conflict"}:
        expected = _parse_bool_text(value)
        if expected == "invalid":
            return False, "has_conflict 只支持 true/false"
        actual = bool(item.get("has_conflict"))
        return actual == expected, None

    if key in {"merged_in_target", "merged"}:
        expected = _parse_merged_in_target_value(value, allow_skipped=True)
        if expected == "invalid":
            return False, "merged_in_target 只支持 true/false/none/skipped"
        actual = item.get("merged_in_target")
        if item.get("is_merge_commit") or actual == "skipped":
            actual = "skipped"
        if isinstance(actual, str):
            actual_text = actual.strip().lower()
        else:
            actual_text = actual
        if expected == "skipped":
            return actual_text == "skipped", None
        return actual_text == expected, None

    if key in {"commit", "sha"}:
        commit_id = str(item.get("commit") or item.get("input_commit") or "")
        return commit_id.lower().startswith(value.lower()), None

    return False, f"不支持的搜索字段: {key}"


def _filter_commit_items(commit_items: list, query: str):
    filtered = []
    for item in commit_items:
        matched, error = _match_commit_item_with_query(item, query)
        if error:
            return None, error
        if matched:
            filtered.append(item)
    return filtered, None


def _parse_index_ranges(text: str, max_index: int):
    if max_index <= 0:
        return [], "当前没有可操作的提交"
    if not text or not text.strip():
        return [], "请输入索引或区间"

    selected = set()
    parts = [part.strip() for part in text.split(",") if part.strip()]
    if not parts:
        return [], "请输入索引或区间"
    for part in parts:
        if "-" in part:
            left, right = part.split("-", 1)
            left = left.strip()
            right = right.strip()
            if not left.isdigit() or not right.isdigit():
                return [], f"区间格式无效: {part}"
            start = int(left)
            end = int(right)
            if start > end:
                start, end = end, start
            if start < 0 or end >= max_index:
                return [], f"索引超出范围: {part}"
            for idx in range(start, end + 1):
                selected.add(idx)
        else:
            if not part.isdigit():
                return [], f"索引格式无效: {part}"
            idx = int(part)
            if idx < 0 or idx >= max_index:
                return [], f"索引超出范围: {part}"
            selected.add(idx)
    return sorted(selected), None


def _build_filtered_report_path(config_path: str):
    if not isinstance(config_path, str) or not config_path:
        return ""
    suffix = ".report.yml"
    if config_path.endswith(suffix):
        prefix = config_path[: -len(suffix)]
    else:
        prefix = config_path
    candidate = f"{prefix}.filtered.report.yml"
    if not os.path.exists(candidate):
        return candidate
    index = 1
    while True:
        candidate = f"{prefix}.filtered.{index}.report.yml"
        if not os.path.exists(candidate):
            return candidate
        index += 1


def _save_filtered_report_config(config_path: str, config: dict, commit_items: list):
    output_path = _build_filtered_report_path(config_path)
    if not output_path:
        return ""
    new_config = dict(config or {})
    new_config["commits"] = commit_items
    with open(output_path, "w", encoding="utf-8") as f:
        yaml.safe_dump(new_config, f, allow_unicode=True, sort_keys=False)
    return output_path


def _interactive_items_changed(original_items: list, working_items: list) -> bool:
    return original_items != working_items


def _interactive_adjust_merged_in_target(config_path: str, config: dict, commit_items: list):
    if not isinstance(commit_items, list) or not commit_items:
        return "continue"
    original_items = copy.deepcopy(commit_items)
    working_items = copy.deepcopy(commit_items)
    print("")
    print("进入交互模式：可修改 merged_in_target（true/false/none）。")
    print("说明：merged_in_target=skipped 代表 merge commit，且不可修改。")
    print("输入格式：<索引> <值>，例如：0 true")
    print("输入 l 重新列表，输入 c 继续执行，输入 q 退出。")
    print("输入 save 或 w 仅保存当前筛选结果并退出。")
    print("输入 search 进入筛选；示例：search> has_conflict=true 或 search> title: drm/hisilicon/hibmc")
    print("输入 d 或 delete 删除当前列表条目；示例：delete> 0-10, 18,19")
    print("输入 reset 恢复到原始列表。")
    while True:
        rows = []
        for idx, item in enumerate(working_items):
            if not isinstance(item, dict):
                display_item = {"commit": item}
            else:
                display_item = item
            display_commit = display_item.get("commit") or display_item.get("input_commit") or ""
            if isinstance(display_commit, str) and _looks_like_commit_sha(display_commit):
                display_commit = display_commit[:12]
            display_merged_status = (
                "skipped"
                if display_item.get("is_merge_commit") or display_item.get("merged_in_target") == "skipped"
                else display_item.get("merged_in_target")
            )
            rows.append([
                idx,
                display_commit,
                (display_item.get("commit_title") or "")[:60],
                display_merged_status,
                display_item.get("has_conflict"),
            ])
        print("")
        print(f"当前显示提交数: {len(working_items)} / 原始总数: {len(original_items)}")
        print(tabulate(rows, headers=["idx", "commit", "title", "merged_in_target", "has_conflict"], tablefmt="github"))
        cmd = _interactive_input("merged_in_target> ").strip()
        if not cmd or cmd.lower() in {"l", "list", "ls"}:
            continue
        if cmd.lower() in {"c", "continue"}:
            commit_items[:] = working_items
            config["commits"] = working_items
            if _interactive_items_changed(original_items, working_items):
                filtered_path = _save_filtered_report_config(config_path, config, working_items)
                if filtered_path:
                    config["_interactive_output_report_path"] = filtered_path
                    print(f"已保存筛选配置: {filtered_path}（commits={len(working_items)}）")
            else:
                print("未检测到筛选变更，继续使用原配置。")
            return "continue"
        if cmd.lower() in {"save", "w"}:
            commit_items[:] = working_items
            config["commits"] = working_items
            if _interactive_items_changed(original_items, working_items):
                filtered_path = _save_filtered_report_config(config_path, config, working_items)
                if filtered_path:
                    config["_interactive_output_report_path"] = filtered_path
                    print(f"已保存筛选配置: {filtered_path}（commits={len(working_items)}）")
            else:
                print("未检测到筛选变更，未生成新配置文件。")
            return "save"
        if cmd.lower() in {"q", "quit", "exit"}:
            return "quit"
        if cmd.lower() in {"search", "s"}:
            query = _interactive_input("search> ").strip()
            filtered_items, search_error = _filter_commit_items(working_items, query)
            if search_error:
                print(f"搜索条件错误：{search_error}")
                continue
            working_items = filtered_items
            print(f"筛选完成：命中 {len(working_items)} 条")
            continue
        if cmd.lower() in {"delete", "d"}:
            delete_expr = _interactive_input("delete> ").strip()
            delete_indexes, delete_error = _parse_index_ranges(delete_expr, len(working_items))
            if delete_error:
                print(f"删除参数错误：{delete_error}")
                continue
            delete_set = set(delete_indexes)
            working_items = [item for idx, item in enumerate(working_items) if idx not in delete_set]
            print(f"删除完成：已删除 {len(delete_set)} 条，剩余 {len(working_items)} 条")
            continue
        if cmd.lower() in {"reset", "r"}:
            working_items = list(original_items)
            print(f"已恢复原始列表，共 {len(working_items)} 条")
            continue
        parts = cmd.split(None, 1)
        if len(parts) != 2:
            print("输入格式错误：请使用 <索引> <值>（true/false/none），或输入 search/delete/reset/save/c/q")
            continue
        try:
            idx = int(parts[0])
        except ValueError:
            print("索引必须是数字")
            continue
        if idx < 0 or idx >= len(working_items):
            print("索引超出范围")
            continue
        merged_value = _parse_merged_in_target_value(parts[1])
        if merged_value == "invalid":
            print("值无效：可用 true/false/none")
            continue
        item = working_items[idx]
        if not isinstance(item, dict):
            item = {"commit": item}
            working_items[idx] = item
        if item.get("is_merge_commit") or item.get("merged_in_target") == "skipped":
            print(f"idx={idx} 是 merge commit，状态为 skipped，不允许修改 merged_in_target")
            continue
        item["merged_in_target"] = merged_value
        item["merged_check_error"] = None
        print(f"已更新 idx={idx} merged_in_target={merged_value}")

def _looks_like_commit_sha(value: str) -> bool:
    if not isinstance(value, str):
        return False
    candidate = value.strip().lower()
    if len(candidate) < 7 or len(candidate) > 40:
        return False
    return all(ch in "0123456789abcdef" for ch in candidate)

def _write_commit_patch_file(commit_id: str, project_dir: str):
    repo = git.Repo(project_dir)
    full_sha = repo.git.rev_parse(commit_id)
    commit_obj = repo.commit(full_sha)
    if len(commit_obj.parents) > 1:
        logger.info(
            "[backport-batch] merge commit 跳过生成原始补丁: commit=%s",
            full_sha,
        )
        return full_sha, ""
    clone_dir = os.path.dirname(project_dir.rstrip("/"))
    patch_path = os.path.join(clone_dir, f"commit_patch_{full_sha}.patch")
    try:
        # 使用 format-patch 生成标准 mbox 格式，便于复用统一的 commit message 解析逻辑
        # 输出形态示例：From/From:/Date:/Subject: ... --- diff --git ...
        patch_content = repo.git.format_patch(
            "-1",
            "--stdout",
            "--no-signature",
            full_sha,
        )
        if patch_content and not patch_content.endswith("\n"):
            # Ensure the patch ends with a newline to avoid "corrupt patch" errors.
            patch_content += "\n"
        with open(patch_path, "w", encoding="utf-8") as f:
            f.write(patch_content)
    except Exception as e:
        raise ValueError(f"生成原始补丁文件失败: {e}")
    return full_sha, patch_path

def _ensure_clean_and_checkout(repo: git.Repo, branch_name: str):
    if repo.is_dirty(untracked_files=True):
        repo.git.reset("--hard")
        repo.git.clean("-fdx")
    try:
        current_branch = repo.active_branch.name if not repo.head.is_detached else None
        if current_branch != branch_name:
            repo.git.checkout(branch_name)
    except Exception as e:
        try:
            if "origin" in [r.name for r in repo.remotes]:
                repo.remotes.origin.fetch()
                origin_ref = f"origin/{branch_name}"
                if origin_ref in [r.name for r in repo.remotes.origin.refs]:
                    repo.git.checkout('-f', '-B', branch_name, origin_ref)
                    return
        except Exception as fetch_error:
            raise ValueError(f"同步远程分支失败: {fetch_error}")
        raise ValueError(f"切换目标分支失败: {e}")

def _abort_cherry_pick(repo: git.Repo) -> None:
    try:
        repo.git.cherry_pick("--abort")
    except Exception:
        pass


def _reset_hard_and_clean(repo: git.Repo, head_sha: str | None = None) -> None:
    try:
        if head_sha:
            repo.git.reset("--hard", head_sha)
        else:
            repo.git.reset("--hard")
        repo.git.clean("-fdx")
    except Exception:
        pass


def _ensure_commit_available_for_cherry_pick(
    target_repo: git.Repo,
    project_dir: str,
    commit_sha: str,
) -> None:
    if os.path.abspath(target_repo.working_dir.rstrip("/")) == os.path.abspath(project_dir.rstrip("/")):
        return
    try:
        target_repo.commit(commit_sha)
        return
    except Exception:
        pass

    upstream_url = os.path.abspath(project_dir.rstrip("/"))
    remote_name = "upstream"
    upstream_remote = next(
        (remote for remote in target_repo.remotes if remote.name == remote_name),
        None,
    )
    if upstream_remote is None:
        upstream_remote = target_repo.create_remote(remote_name, upstream_url)
    upstream_remote.fetch(commit_sha)


def _is_commit_merged_in_target(target_path: str, target_branch: str, commit_sha: str):
    try:
        target_repo = git.Repo(target_path)
        _ensure_clean_and_checkout(target_repo, target_branch)
        target_repo.git.merge_base("--is-ancestor", commit_sha, target_branch)
        return True, None
    except git.exc.GitCommandError as e:
        # 非祖先或分支不存在都会抛错，按未合入处理
        return False, str(e)
    except Exception as e:
        return False, str(e)

def _is_patch_applied_in_target(target_path: str, target_branch: str, patch_path: str):
    if not patch_path or not os.path.exists(patch_path):
        return False, "patch_path missing"
    try:
        target_repo = git.Repo(target_path)
        _ensure_clean_and_checkout(target_repo, target_branch)
        target_repo.git.apply("--check", "--reverse", patch_path)
        return True, None
    except git.exc.GitCommandError as e:
        return False, str(e)
    except Exception as e:
        return False, str(e)

def _check_conflict_with_apply_or_cherrypick(
    target_path: str,
    target_branch: str,
    commit_sha: str,
    patch_path: str,
    project_dir: str,
):
    target_repo = git.Repo(target_path)
    _ensure_clean_and_checkout(target_repo, target_branch)
    upstream_repo = git.Repo(project_dir)
    commit_obj = upstream_repo.commit(commit_sha)
    is_merge_commit = len(commit_obj.parents) > 1
    if is_merge_commit:
        return False, "merge-commit-skipped", None
    apply_error = None
    # 优先用 git apply --check
    try:
        target_repo.git.apply("--check", patch_path)
        return False, "apply", None
    except git.exc.GitCommandError as e:
        apply_error = str(e)

    # 再尝试 cherry-pick
    original_head = target_repo.head.commit.hexsha
    try:
        _ensure_commit_available_for_cherry_pick(target_repo, project_dir, commit_sha)
        target_repo.git.cherry_pick(commit_sha)
        _reset_hard_and_clean(target_repo, original_head)
        return False, "cherry-pick", None
    except git.exc.GitCommandError as e:
        _abort_cherry_pick(target_repo)
        _reset_hard_and_clean(target_repo, original_head)
        if apply_error:
            return True, "cherry-pick", f"{apply_error}; cherry-pick error: {e}"
        return True, "cherry-pick", f"cherry-pick error: {e}"
    except Exception as e:
        _abort_cherry_pick(target_repo)
        _reset_hard_and_clean(target_repo, original_head)
        if apply_error:
            return True, "cherry-pick", f"{apply_error}; cherry-pick error: {e}"
        return True, "cherry-pick", f"cherry-pick error: {e}"

def _apply_patch_to_target_repo(
    target_path: str,
    target_branch: str,
    commit_sha: str,
    project_dir: str,
    commit_message: str,
    signer_name: str = "",
    signer_email: str = "",
):
    target_repo = git.Repo(target_path)
    _ensure_clean_and_checkout(target_repo, target_branch)
    original_head = target_repo.head.commit.hexsha
    try:
        upstream_repo = git.Repo(project_dir)
        commit_obj = upstream_repo.commit(commit_sha)
        is_merge_commit = len(commit_obj.parents) > 1
        if is_merge_commit:
            return {"status": "skipped", "error": "merge commit 已跳过"}
        _ensure_commit_available_for_cherry_pick(target_repo, project_dir, commit_sha)
        target_repo.git.cherry_pick("--no-commit", commit_sha)
        if not target_repo.is_dirty(index=True, working_tree=True, untracked_files=False):
            return {"status": "skipped", "error": "补丁未产生变更（可能已等效存在）"}
        commit_args = ["-m", commit_message, "-s"]
        use_signer_identity = bool(signer_name and signer_email)
        git_env = {
            "GIT_AUTHOR_NAME": commit_obj.author.name,
            "GIT_AUTHOR_EMAIL": commit_obj.author.email,
        }
        if use_signer_identity:
            git_env["GIT_COMMITTER_NAME"] = signer_name
            git_env["GIT_COMMITTER_EMAIL"] = signer_email
        with target_repo.git.custom_environment(**git_env):
            target_repo.git.commit(*commit_args)
        return {"status": "success", "commit": target_repo.head.commit.hexsha}
    except Exception as e:
        _abort_cherry_pick(target_repo)
        _reset_hard_and_clean(target_repo, original_head)
        return {"status": "failed", "error": str(e)}

def _resolve_sorted_backport_items(commit_items, is_report_config, base_project_dir, base_config, args):
    try:
        return resolve_sorted_backport_items(
            commit_items,
            is_report_config,
            base_project_dir,
            base_config,
            args,
        )
    except ValueError as e:
        raise ValueError(f"{e}（请先克隆源代码仓到 project_dir）")


def _append_sort_error_report_item(report_items, item, error_message):
    commit_id, commit_title, item_config = _extract_commit_item(item)
    tag = item_config.get("tag") or commit_id or commit_title or ""
    logger.info("[backport-batch] 排序错误: tag=%s, error=%s", tag, error_message)
    report_items.append({
        "commit": None,
        "input_commit": commit_id,
        "commit_title": commit_title,
        "committed_datetime": None,
        "target_branch": item_config.get("target_branch") or item_config.get("target_release"),
        "status": "failed",
        "has_conflict": None,
        "conflict_check_method": None,
        "conflict_check_error": error_message,
        "original_patch_path": None,
        "backported_patch_path": None,
        "patch_path": None,
        "error": error_message,
    })


def _resolve_target_branch(item_config, base_config, default_target_branch):
    return (
        item_config.get("target_branch")
        or item_config.get("target_release")
        or base_config.get("target_branch")
        or base_config.get("target_release")
        or default_target_branch
    )


def _append_remaining_pending_items(
    report_items,
    remaining_items,
    is_report_config,
    base_config,
    default_target_branch,
):
    for remaining in remaining_items:
        remaining_commit_id = remaining.get("commit")
        remaining_input_commit = remaining.get("input_commit")
        remaining_commit_title = remaining.get("commit_title")
        remaining_git_describe = remaining.get("git_describe")
        remaining_item_config = (
            remaining if is_report_config else remaining.get("item_config", {})
        )
        remaining_committed_datetime = remaining.get("committed_datetime")
        remaining_target_branch = _resolve_target_branch(
            remaining_item_config, base_config, default_target_branch
        )
        remaining_original_patch_path = (
            remaining_item_config.get("original_patch_path")
            or remaining_item_config.get("patch_path")
            or ""
        )
        # report 重跑场景下，尽量保留历史检测状态，避免被 pending 覆盖为 null
        remaining_merged_in_target = remaining_item_config.get("merged_in_target")
        remaining_merged_check_error = remaining_item_config.get("merged_check_error")
        remaining_has_conflict = remaining_item_config.get("has_conflict")
        remaining_conflict_check_method = remaining_item_config.get("conflict_check_method")
        remaining_conflict_check_error = remaining_item_config.get("conflict_check_error")
        remaining_status = "pending"
        if is_report_config and remaining_item_config.get("status"):
            remaining_status = remaining_item_config.get("status")
        pending_item = {
            "commit": remaining_commit_id,
            "input_commit": remaining_input_commit,
            "commit_title": remaining_commit_title,
            "committed_datetime": remaining_committed_datetime,
            "git_describe": remaining_git_describe,
            "target_branch": remaining_target_branch,
            "status": remaining_status,
            "merged_in_target": remaining_merged_in_target,
            "merged_check_error": remaining_merged_check_error,
            "is_merge_commit": remaining_item_config.get("is_merge_commit"),
            "has_conflict": remaining_has_conflict,
            "conflict_check_method": remaining_conflict_check_method,
            "conflict_check_error": remaining_conflict_check_error,
            "original_patch_path": remaining_original_patch_path,
            "backported_patch_path": None,
            "patch_path": remaining_original_patch_path,
            "error": None,
        }
        if remaining_item_config.get("backport_engine"):
            pending_item["backport_engine"] = remaining_item_config["backport_engine"]
        report_items.append(pending_item)


def _build_batch_summary_result(
    tag,
    commit_id,
    target_branch,
    fixed_commit,
    merged_in_target,
    merged_check_error,
    has_conflict,
    conflict_check_method,
    conflict_check_error,
    backport_result,
):
    empty_patch = bool(backport_result.get("empty_patch"))
    equivalent_exists = bool(backport_result.get("equivalent_exists"))
    empty_and_equivalent = empty_patch and equivalent_exists
    short_note = (
        i18n("补丁实际上已应用（或等效存在），无需继续执行后续步骤")
        if empty_and_equivalent and backport_result.get("status") == "success"
        else ""
    )
    display_status = (
        i18n("无需执行（补丁已存在）")
        if empty_and_equivalent and backport_result.get("status") == "success"
        else (i18n("成功") if backport_result.get("status") == "success" else i18n("需要调整"))
    )
    result = {
        i18n("补丁ID"): tag or commit_id,
        i18n("目标分支"): target_branch,
        i18n("适配状态"): display_status,
        i18n("冲突点"): "" if empty_and_equivalent else (backport_result.get("backported_patch_path") or backport_result.get("original_patch_path", "")),
        i18n("建议调整文件"): "N/A" if backport_result.get("status") == "success" else "",
    }
    if short_note:
        # 兼容仅展示 error 字段的消费方，给出同样清晰的结论
        result["error"] = short_note
    result["details"] = {
        "action": "backport-batch",
        "tag": tag or commit_id,
        "target_branch": target_branch,
        "fixed_commit": fixed_commit,
        "merged_in_target": merged_in_target,
        "merged_check_error": merged_check_error,
        "conflict_check": {
            "has_conflict": False if empty_and_equivalent else has_conflict,
            "method": "llm-equivalence" if empty_and_equivalent else conflict_check_method,
            "error": None if empty_and_equivalent else conflict_check_error,
        },
        "original_patch_path": backport_result.get("original_patch_path"),
        "backported_patch_path": backport_result.get("backported_patch_path"),
        "diff_path": backport_result.get("diff_path"),
        "empty_patch": empty_patch,
        "equivalent_exists": equivalent_exists,
        "error": short_note if short_note else backport_result.get("error"),
        "logfile": backport_result.get("logfile"),
        "time_cost": backport_result.get("time_cost"),
        "status": backport_result.get("status"),
    }
    if "cost" in backport_result:
        result["details"]["cost"] = backport_result["cost"]
    if "tokens" in backport_result:
        result["details"]["tokens"] = backport_result["tokens"]
    if "error" in backport_result and not short_note:
        result["details"]["error"] = backport_result["error"]
    return result


def _extract_backport_batch_item_fields(item, is_report_config):
    commit_id = item.get("commit")
    input_commit = item.get("input_commit")
    commit_title = item.get("commit_title")
    item_config = item if is_report_config else item.get("item_config", {})
    committed_datetime = item.get("committed_datetime")
    git_describe = item.get("git_describe")
    is_merge_commit = item_config.get("is_merge_commit")
    tag = item_config.get("tag") or commit_id or input_commit or commit_title
    return {
        "commit_id": commit_id,
        "input_commit": input_commit,
        "commit_title": commit_title,
        "item_config": item_config,
        "committed_datetime": committed_datetime,
        "git_describe": git_describe,
        "is_merge_commit": is_merge_commit,
        "tag": tag,
    }


def _resolve_is_merge_commit(base_project_dir, fixed_commit):
    try:
        upstream_repo = git.Repo(base_project_dir)
        commit_obj = upstream_repo.commit(fixed_commit)
        return len(commit_obj.parents) > 1
    except Exception:
        return None


def _prepare_backport_patch_and_commit(
    *,
    is_report_config,
    generate_missing_patch,
    fixed_commit,
    patch_path,
    is_merge_commit,
    base_project_dir,
):
    if (not is_report_config) or (generate_missing_patch and not patch_path):
        try:
            logger.info("[backport-batch] 生成补丁: commit=%s", fixed_commit)
            full_sha, patch_path = _write_commit_patch_file(fixed_commit, base_project_dir)
            fixed_commit = full_sha
            if is_merge_commit is None:
                is_merge_commit = _resolve_is_merge_commit(base_project_dir, fixed_commit)
            logger.info("[backport-batch] 补丁生成成功: commit=%s, patch=%s", fixed_commit, patch_path)
            return fixed_commit, patch_path, is_merge_commit, False
        except Exception as e:
            logger.info("[backport-batch] 补丁生成失败: commit=%s, error=%s", fixed_commit, e)
            return fixed_commit, patch_path, is_merge_commit, True
    if is_merge_commit is None and fixed_commit:
        is_merge_commit = _resolve_is_merge_commit(base_project_dir, fixed_commit)
    return fixed_commit, patch_path, is_merge_commit, False


def _resolve_merge_and_conflict_status(
    *,
    is_report_config,
    force_recheck,
    base_target_path,
    target_branch,
    fixed_commit,
    patch_path,
    base_project_dir,
    is_merge_commit,
    tag,
    commit_id,
    merged_in_target,
    merged_check_error,
    has_conflict,
    conflict_check_method,
    conflict_check_error,
):
    if not fixed_commit:
        return (
            merged_in_target,
            merged_check_error,
            has_conflict,
            conflict_check_method,
            conflict_check_error,
        )

    if is_merge_commit:
        logger.info(
            "[backport-batch] merge commit 跳过合入/冲突检测: tag=%s, commit=%s, branch=%s",
            tag or commit_id,
            fixed_commit,
            target_branch,
        )
        return (
            "skipped",
            "merge commit skipped",
            False,
            "merge-commit-skipped",
            None,
        )

    use_config_merged = is_report_config and merged_in_target is not None and not force_recheck
    if use_config_merged:
        logger.info(
            "[backport-batch] 使用 report 配置中的 merged_in_target: tag=%s, merged_in_target=%s",
            tag or commit_id,
            merged_in_target,
        )
        if merged_in_target:
            has_conflict, conflict_check_method, conflict_check_error = False, "merged", None
        return (
            merged_in_target,
            merged_check_error,
            has_conflict,
            conflict_check_method,
            conflict_check_error,
        )

    merged_in_target, merged_check_error = _is_commit_merged_in_target(
        base_target_path, target_branch, fixed_commit
    )
    if merged_in_target:
        logger.info(
            "[backport-batch] 目标分支已包含该提交，跳过冲突检测: tag=%s, commit=%s, branch=%s",
            tag or commit_id,
            fixed_commit,
            target_branch,
        )
        has_conflict, conflict_check_method, conflict_check_error = False, "merged", None
    else:
        patch_applied = False
        patch_check_error = None
        if patch_path:
            patch_applied, patch_check_error = _is_patch_applied_in_target(
                base_target_path, target_branch, patch_path
            )
        if patch_applied:
            merged_in_target = True
            merged_check_error = None
            has_conflict, conflict_check_method, conflict_check_error = False, "patch-reverse", None
            logger.info(
                "[backport-batch] 目标分支已包含该补丁内容，跳过冲突检测: tag=%s, commit=%s, branch=%s",
                tag or commit_id,
                fixed_commit,
                target_branch,
            )
        else:
            if patch_check_error:
                if merged_check_error:
                    merged_check_error = f"{merged_check_error}; patch-reverse-check: {patch_check_error}"
                else:
                    merged_check_error = f"patch-reverse-check: {patch_check_error}"
            logger.info(
                "[backport-batch] 目标分支合入检查未命中: tag=%s, commit=%s, branch=%s, error=%s",
                tag or commit_id,
                fixed_commit,
                target_branch,
                merged_check_error,
            )

    # Raw configs need an initial conflict check. Report configs only recheck
    # when the caller explicitly enters stop-at-first-conflict detection mode.
    if ((not is_report_config) or force_recheck) and (not merged_in_target):
        has_conflict, conflict_check_method, conflict_check_error = _check_conflict_with_apply_or_cherrypick(
            base_target_path,
            target_branch,
            fixed_commit,
            patch_path,
            base_project_dir,
        )
        logger.info(
            "[backport-batch] 冲突检测: tag=%s, has_conflict=%s, method=%s, error=%s",
            tag or commit_id,
            has_conflict,
            conflict_check_method,
            conflict_check_error,
        )
    elif (not is_report_config) or force_recheck:
        logger.info(
            "[backport-batch] 已判定目标分支包含补丁，跳过二次冲突检测: tag=%s, commit=%s, branch=%s",
            tag or commit_id,
            fixed_commit,
            target_branch,
        )

    return (
        merged_in_target,
        merged_check_error,
        has_conflict,
        conflict_check_method,
        conflict_check_error,
    )


def _build_backport_runtime_config(
    *,
    item_config,
    base_config,
    base_project_dir,
    base_target_path,
    fixed_commit,
    target_branch,
    tag,
    commit_id,
    args,
):
    base_dataset_dir = (
        item_config.get("patch_dataset_dir")
        or base_config.get("patch_dataset_dir")
        or args.patch_dataset_dir
        or os.path.join(os.path.expanduser("~"), "backports/patch_dataset")
    )
    dataset_dir = os.path.join(base_dataset_dir, str(tag or commit_id))
    logger.info("[backport-batch] 数据集目录: %s", dataset_dir)

    config_dict = {
        "project": base_config.get("project", "linux"),
        "project_url": base_config.get("project_url", ""),
        "project_dir": base_project_dir,
        "target_path": base_target_path,
        "new_patch": fixed_commit,
        "target_release": target_branch,
        "openai_key": (
            item_config.get("openai_key")
            or item_config.get("api_key")
            or base_config.get("openai_key")
            or base_config.get("api_key")
            or args.api_key
        ),
        "llm_provider": item_config.get("llm_provider") or base_config.get("llm_provider") or args.llm_provider,
        "llm_base_url": item_config.get("llm_base_url") or base_config.get("llm_base_url") or getattr(args, 'llm_base_url', None),
        "llm_model_name": item_config.get("llm_model_name") or base_config.get("llm_model_name") or getattr(args, 'llm_model_name', None),
        "tag": tag or commit_id,
        "patch_dataset_dir": dataset_dir,
        "skip_cherry_pick": True,
        "signer_name": item_config.get("signer_name") or base_config.get("signer_name") or getattr(args, "signer_name", None),
        "signer_email": item_config.get("signer_email") or base_config.get("signer_email") or getattr(args, "signer_email", None),
        "commit_message_template": (
            str(getattr(args, "commit_message_template", "") or "").strip()
            or str(item_config.get("commit_message_template") or "").strip()
            or str(base_config.get("commit_message_template") or "").strip()
            or DEFAULT_COMMIT_MESSAGE_TEMPLATE
        ),
        "linux_repo_path": (
            str(getattr(args, "linux_repo_path", "") or "").strip()
            or str(item_config.get("linux_repo_path") or "").strip()
            or str(base_config.get("linux_repo_path") or "").strip()
            or DEFAULT_LINUX_REPO_PATH
        ),
        "commit_message_source": _resolve_commit_message_source(args, item_config, base_config),
        "backport_engine": _resolve_backport_engine(args, item_config, base_config),
        "format_mode": (
            item_config.get("format_mode")
            or base_config.get("format_mode")
            or getattr(args, "format_mode", None)
            or "full"
        ),
    }
    if item_config.get("error_message") or base_config.get("error_message") or args.error_message:
        config_dict["error_message"] = (
            item_config.get("error_message") or base_config.get("error_message") or args.error_message
        )
    if item_config.get("sanitizer") or base_config.get("sanitizer") or args.sanitizer:
        config_dict["sanitizer"] = item_config.get("sanitizer") or base_config.get("sanitizer") or args.sanitizer
    return config_dict


def _run_mystique_from_config(config_dict: dict, debug_mode: bool = False) -> dict:
    """运行 Mystique，并转换为 backport-batch 统一使用的结果结构。"""
    mystique_src = os.path.join(os.path.dirname(__file__), "mystique", "src")
    if mystique_src not in sys.path:
        sys.path.insert(0, mystique_src)
    import config as mystique_config
    import main as mystique_main

    mystique_config.configure_llm(
        provider=config_dict.get("llm_provider"),
        api_key=config_dict.get("openai_key"),
        base_url=config_dict.get("llm_base_url"),
        model_name=config_dict.get("llm_model_name"),
    )
    mystique_config.configure_format_normalization(config_dict.get("format_mode"))

    start_time = time.time()
    results = mystique_main.main_from_repo(
        project_dir=config_dict["project_dir"],
        target_path=config_dict["target_path"],
        new_patch=config_dict["new_patch"],
        target_release=config_dict["target_release"],
        signatures=None,
        output=config_dict.get("patch_dataset_dir"),
        cve_id=str(config_dict.get("tag") or config_dict["new_patch"]),
        skip_cherry_pick=bool(config_dict.get("skip_cherry_pick", False)),
        debug=debug_mode,
    )
    original_patch_path = next(
        (item.get("original_patch_path") for item in results if item.get("original_patch_path")),
        None,
    )
    backported_patch_path = next(
        (item.get("backported_patch_path") for item in results if item.get("backported_patch_path")),
        None,
    )
    logfile = next(
        (item.get("logfile") for item in results if item.get("logfile")),
        None,
    )
    empty_patch = bool(results) and all(
        item.get("status") == "need_not_ported" for item in results
    )
    status = "success" if backported_patch_path or empty_patch else "failed"
    result = {
        "status": status,
        "original_patch_path": original_patch_path,
        "backported_patch_path": backported_patch_path,
        "diff_path": backported_patch_path,
        "empty_patch": empty_patch,
        "equivalent_exists": empty_patch,
        "logfile": logfile,
        "time_cost": int(time.time() - start_time),
    }
    if status == "failed":
        result["error"] = "Mystique 未生成回移植补丁"
    return result


def _run_selected_backport_engine(config_dict: dict, debug_mode: bool = False) -> dict:
    """根据 batch 配置选择 PortGPT 或 Mystique，并返回统一结果。"""
    engine = str(config_dict.get("backport_engine") or "portgpt").strip().lower()
    logger.info("[backport-batch] 使用回移植引擎: %s", engine)
    if engine == "portgpt":
        return run_backport_from_config(config_dict, debug_mode=debug_mode)
    if engine == "mystique":
        return _run_mystique_from_config(config_dict, debug_mode=debug_mode)
    raise ValueError(f"不支持的 backport_engine: {engine!r}")


def _execute_backport_batch_action(
    *,
    is_report_config,
    is_merge_commit,
    merged_in_target,
    has_conflict,
    conflict_check_method,
    conflict_check_error,
    item_config,
    tag,
    commit_id,
    target_branch,
    fixed_commit,
    patch_path,
    config_dict,
    base_target_path,
    base_project_dir,
    args,
):
    did_backport = False
    refreshed_status = None
    if is_merge_commit:
        logger.info(
            "[backport-batch] merge commit 跳过处理: tag=%s, commit=%s, branch=%s",
            tag or commit_id,
            fixed_commit,
            target_branch,
        )
        return {
            "status": "skipped",
            "original_patch_path": item_config.get("original_patch_path", "") or patch_path,
            "backported_patch_path": item_config.get("backported_patch_path", ""),
            "diff_path": item_config.get("diff_path", ""),
            "logfile": item_config.get("logfile", ""),
            "time_cost": 0,
            "error": "merge commit 已跳过",
        }, did_backport, refreshed_status

    if is_report_config and getattr(args, "stop_at_first_conflict", False):
        logger.info(
            "[backport-batch] report 检测模式，仅更新状态不应用补丁: tag=%s, commit=%s, has_conflict=%s",
            tag or commit_id,
            fixed_commit,
            has_conflict,
        )
        return {
            "status": "success",
            "original_patch_path": item_config.get("original_patch_path", "") or patch_path,
            "backported_patch_path": item_config.get("backported_patch_path", "") if has_conflict else "",
            "diff_path": item_config.get("diff_path", ""),
            "logfile": item_config.get("logfile", ""),
            "time_cost": 0,
        }, did_backport, refreshed_status

    if is_report_config:
        if merged_in_target:
            logger.info(
                "[backport-batch] 目标分支已包含该提交，跳过应用补丁: tag=%s, branch=%s",
                tag or commit_id,
                target_branch,
            )
            return {
                "status": "skipped",
                "original_patch_path": item_config.get("original_patch_path", ""),
                "backported_patch_path": item_config.get("backported_patch_path", ""),
                "diff_path": item_config.get("diff_path", ""),
                "logfile": item_config.get("logfile", ""),
                "time_cost": 0,
                "error": "目标分支已包含该提交",
            }, did_backport, refreshed_status

        if has_conflict:
            # report 场景下，冲突可能因前置补丁已应用而发生变化，进入回移植前先实时复检一次
            latest_merged_in_target, latest_merged_check_error = _is_commit_merged_in_target(
                base_target_path, target_branch, fixed_commit
            )
            latest_has_conflict = has_conflict
            latest_conflict_check_method = conflict_check_method
            latest_conflict_check_error = conflict_check_error

            if latest_merged_in_target:
                latest_has_conflict = False
                latest_conflict_check_method = "merged"
                latest_conflict_check_error = None
            else:
                patch_applied = False
                patch_check_error = None
                if patch_path:
                    patch_applied, patch_check_error = _is_patch_applied_in_target(
                        base_target_path, target_branch, patch_path
                    )
                if patch_applied:
                    latest_merged_in_target = True
                    latest_merged_check_error = None
                    latest_has_conflict = False
                    latest_conflict_check_method = "patch-reverse"
                    latest_conflict_check_error = None
                else:
                    if patch_check_error:
                        if latest_merged_check_error:
                            latest_merged_check_error = (
                                f"{latest_merged_check_error}; patch-reverse-check: {patch_check_error}"
                            )
                        else:
                            latest_merged_check_error = f"patch-reverse-check: {patch_check_error}"
                    (
                        latest_has_conflict,
                        latest_conflict_check_method,
                        latest_conflict_check_error,
                    ) = _check_conflict_with_apply_or_cherrypick(
                        base_target_path,
                        target_branch,
                        fixed_commit,
                        patch_path,
                        base_project_dir,
                    )

            refreshed_status = {
                "merged_in_target": latest_merged_in_target,
                "merged_check_error": latest_merged_check_error,
                "has_conflict": latest_has_conflict,
                "conflict_check_method": latest_conflict_check_method,
                "conflict_check_error": latest_conflict_check_error,
            }
            logger.info(
                "[backport-batch] 冲突复检结果: tag=%s, merged_in_target=%s, has_conflict=%s, method=%s, error=%s",
                tag or commit_id,
                latest_merged_in_target,
                latest_has_conflict,
                latest_conflict_check_method,
                latest_conflict_check_error,
            )

            if not latest_has_conflict:
                if latest_merged_in_target:
                    logger.info(
                        "[backport-batch] 冲突复检后发现补丁已生效，跳过回移植: tag=%s, branch=%s",
                        tag or commit_id,
                        target_branch,
                    )
                    return {
                        "status": "skipped",
                        "original_patch_path": item_config.get("original_patch_path", "") or patch_path,
                        "backported_patch_path": "",
                        "diff_path": "",
                        "logfile": "",
                        "time_cost": 0,
                        "error": "冲突复检后发现目标分支已包含该补丁",
                    }, did_backport, refreshed_status

                logger.info(
                    "[backport-batch] 冲突复检后已无冲突，直接应用补丁: tag=%s, commit=%s, target_branch=%s",
                    tag or commit_id,
                    fixed_commit,
                    target_branch,
                )
                commit_message_source = _infer_original_patch_path(item_config) or patch_path
                commit_message_preview = _resolve_commit_message_fields(
                    patch_path=commit_message_source,
                    openeuler_commit_id=str(commit_id or fixed_commit or ""),
                    item_config=item_config,
                    base_config=config_dict,
                    args=args,
                )
                apply_result = _apply_patch_to_target_repo(
                    base_target_path,
                    target_branch,
                    fixed_commit,
                    base_project_dir,
                    commit_message_preview["commit_message"],
                    signer_name=str(getattr(args, "signer_name", "") or config_dict.get("signer_name") or "").strip(),
                    signer_email=str(getattr(args, "signer_email", "") or config_dict.get("signer_email") or "").strip(),
                )
                result = {
                    "status": apply_result.get("status"),
                    "original_patch_path": item_config.get("original_patch_path", "") or patch_path,
                    "backported_patch_path": "",
                    "diff_path": "",
                    "logfile": "",
                    "time_cost": 0,
                    "commit_message_preview": commit_message_preview.get("commit_message_preview", ""),
                    "commit_message_context": commit_message_preview.get("commit_message_context", {}),
                    "source_detection": commit_message_preview.get("source_detection", {}),
                    "commit_message_warnings": commit_message_preview.get("commit_message_warnings", []),
                }
                if apply_result.get("error"):
                    result["error"] = apply_result["error"]
                return result, did_backport, refreshed_status

            logger.info(
                "[backport-batch] 冲突存在，执行回移植：tag=%s, commit=%s, target_branch=%s",
                tag or commit_id,
                fixed_commit,
                target_branch,
            )
            logger.info("[backport-batch] 回移植配置: %s", config_dict)
            did_backport = True
            return _run_selected_backport_engine(config_dict, debug_mode=args.debug), did_backport, refreshed_status

        logger.info(
            "[backport-batch] 无冲突，直接应用补丁: tag=%s, commit=%s, target_branch=%s",
            tag or commit_id,
            fixed_commit,
            target_branch,
        )
        logger.info("[backport-batch] 应用补丁路径: %s", patch_path)
        if not patch_path:
            return {
                "status": "failed",
                "original_patch_path": item_config.get("original_patch_path", ""),
                "backported_patch_path": item_config.get("backported_patch_path", ""),
                "diff_path": item_config.get("diff_path", ""),
                "logfile": item_config.get("logfile", ""),
                "time_cost": 0,
                "error": "缺少 patch_path，无法应用补丁",
            }, did_backport, refreshed_status
        commit_message_source = _infer_original_patch_path(item_config) or patch_path
        commit_message_preview = _resolve_commit_message_fields(
            patch_path=commit_message_source,
            openeuler_commit_id=str(commit_id or fixed_commit or ""),
            item_config=item_config,
            base_config=config_dict,
            args=args,
        )
        apply_result = _apply_patch_to_target_repo(
            base_target_path,
            target_branch,
            fixed_commit,
            base_project_dir,
            commit_message_preview["commit_message"],
            signer_name=str(getattr(args, "signer_name", "") or config_dict.get("signer_name") or "").strip(),
            signer_email=str(getattr(args, "signer_email", "") or config_dict.get("signer_email") or "").strip(),
        )
        result = {
            "status": apply_result.get("status"),
            "original_patch_path": item_config.get("original_patch_path", "") or patch_path,
            # 无冲突直接应用场景不应继承历史回移植产物路径
            "backported_patch_path": "",
            "diff_path": "",
            "logfile": "",
            "time_cost": 0,
            "commit_message_preview": commit_message_preview.get("commit_message_preview", ""),
            "commit_message_context": commit_message_preview.get("commit_message_context", {}),
            "source_detection": commit_message_preview.get("source_detection", {}),
            "commit_message_warnings": commit_message_preview.get("commit_message_warnings", []),
        }
        if apply_result.get("error"):
            result["error"] = apply_result["error"]
        return result, did_backport, refreshed_status

    logger.info(
        "[backport-batch] 原始配置仅记录结果，不应用补丁: tag=%s, branch=%s",
        tag or commit_id,
        target_branch,
    )
    return {
        "status": "success",
        "original_patch_path": item_config.get("original_patch_path", "") or patch_path,
        "backported_patch_path": "",
        "diff_path": "",
        "logfile": "",
        "time_cost": 0,
    }, did_backport, refreshed_status


def _build_backport_batch_report_item(
    *,
    commit_id,
    input_commit,
    commit_title,
    committed_datetime,
    git_describe,
    target_branch,
    backport_result,
    item_config,
    patch_path,
    merged_in_target,
    merged_check_error,
    is_merge_commit,
    has_conflict,
    conflict_check_method,
    conflict_check_error,
    base_config,
    args,
):
    empty_patch = bool(backport_result.get("empty_patch"))
    equivalent_exists = bool(backport_result.get("equivalent_exists"))
    empty_and_equivalent = empty_patch and equivalent_exists
    short_note = (
        "patch already applied (or equivalent), skip remaining steps"
        if empty_and_equivalent and backport_result.get("status") == "success"
        else ""
    )
    resolved_original_patch = (
        backport_result.get("original_patch_path") or item_config.get("original_patch_path") or patch_path
    )
    resolved_backported_patch = backport_result.get("backported_patch_path")
    effective_merged_in_target = False if empty_and_equivalent else merged_in_target
    effective_has_conflict = False if empty_and_equivalent else has_conflict
    effective_conflict_check_method = "llm-equivalence" if empty_and_equivalent else conflict_check_method
    effective_conflict_check_error = None if empty_and_equivalent else conflict_check_error
    effective_patch_path = (
        ""
        if empty_and_equivalent
        else (resolved_backported_patch if effective_has_conflict else resolved_original_patch)
    )
    message_fields = {
        "commit_message_preview": (
            backport_result.get("commit_message_preview")
            or item_config.get("commit_message_preview")
            or ""
        ),
        "commit_message_context": (
            backport_result.get("commit_message_context")
            or item_config.get("commit_message_context")
            or {}
        ),
        "source_detection": (
            backport_result.get("source_detection")
            or item_config.get("source_detection")
            or {}
        ),
        "commit_message_warnings": (
            backport_result.get("commit_message_warnings")
            or item_config.get("commit_message_warnings")
            or []
        ),
    }
    if not message_fields["commit_message_preview"] and resolved_original_patch:
        message_fields = _build_commit_message_report_fields(
            patch_path=resolved_original_patch,
            openeuler_commit_id=str(commit_id or input_commit or ""),
            item_config=item_config,
            base_config=base_config,
            args=args,
        )
    return {
        "commit": commit_id,
        "input_commit": input_commit,
        "commit_title": commit_title,
        "committed_datetime": committed_datetime,
        "git_describe": git_describe,
        "target_branch": target_branch,
        "backport_engine": _resolve_backport_engine(args, item_config, base_config),
        "status": backport_result.get("status"),
        "merged_in_target": effective_merged_in_target,
        "merged_check_error": merged_check_error,
        "is_merge_commit": is_merge_commit,
        "has_conflict": effective_has_conflict,
        "conflict_check_method": effective_conflict_check_method,
        "conflict_check_error": effective_conflict_check_error,
        "empty_patch": empty_patch,
        "equivalent_exists": equivalent_exists,
        "error": short_note if short_note else backport_result.get("error"),
        "original_patch_path": resolved_original_patch,
        "backported_patch_path": resolved_backported_patch,
        "patch_path": effective_patch_path,
        **message_fields,
    }


def _build_backport_batch_failed_result(
    *,
    did_backport,
    tag,
    commit_id,
    target_branch,
    fixed_commit,
    has_conflict,
    conflict_check_method,
    conflict_check_error,
    error,
):
    if not did_backport:
        return None
    return {
        i18n("补丁ID"): tag or commit_id,
        i18n("目标分支"): target_branch,
        i18n("适配状态"): i18n("需要调整"),
        i18n("冲突点"): "",
        i18n("建议调整文件"): "",
        "details": {
            "action": "backport-batch",
            "tag": tag or commit_id,
            "target_branch": target_branch,
            "fixed_commit": fixed_commit,
            "conflict_check": {
                "has_conflict": has_conflict,
                "method": conflict_check_method,
                "error": conflict_check_error,
            },
            "status": "failed",
            "error": str(error),
        },
    }


def _process_backport_batch_item(
    item,
    is_report_config,
    base_config,
    base_project_dir,
    base_target_path,
    default_target_branch,
    args,
):
    fields = _extract_backport_batch_item_fields(item, is_report_config)
    commit_id = fields["commit_id"]
    input_commit = fields["input_commit"]
    commit_title = fields["commit_title"]
    item_config = fields["item_config"]
    committed_datetime = fields["committed_datetime"]
    git_describe = fields["git_describe"]
    is_merge_commit = fields["is_merge_commit"]
    tag = fields["tag"]

    logger.info(
        "[backport-batch] 处理条目: tag=%s, commit=%s, input_commit=%s, title=%s, time=%s",
        tag,
        commit_id,
        input_commit,
        commit_title,
        committed_datetime,
    )
    target_branch = _resolve_target_branch(item_config, base_config, default_target_branch)
    if not target_branch:
        logger.info("[backport-batch] 目标分支缺失: tag=%s", tag or commit_id)
        return {"skip": True, "did_backport": False}

    fixed_commit = commit_id
    patch_path = item_config.get("patch_path") or item_config.get("original_patch_path") or ""
    fixed_commit, patch_path, is_merge_commit, should_skip = _prepare_backport_patch_and_commit(
        is_report_config=is_report_config,
        generate_missing_patch=bool(is_report_config and getattr(args, "stop_at_first_conflict", False)),
        fixed_commit=fixed_commit,
        patch_path=patch_path,
        is_merge_commit=is_merge_commit,
        base_project_dir=base_project_dir,
    )
    if should_skip:
        return {"skip": True, "did_backport": False}

    merged_in_target = item_config.get("merged_in_target")
    merged_check_error = item_config.get("merged_check_error")
    has_conflict = item_config.get("has_conflict")
    conflict_check_method = item_config.get("conflict_check_method")
    conflict_check_error = item_config.get("conflict_check_error")
    (
        merged_in_target,
        merged_check_error,
        has_conflict,
        conflict_check_method,
        conflict_check_error,
    ) = _resolve_merge_and_conflict_status(
        is_report_config=is_report_config,
        force_recheck=bool(is_report_config and getattr(args, "stop_at_first_conflict", False)),
        base_target_path=base_target_path,
        target_branch=target_branch,
        fixed_commit=fixed_commit,
        patch_path=patch_path,
        base_project_dir=base_project_dir,
        is_merge_commit=is_merge_commit,
        tag=tag,
        commit_id=commit_id,
        merged_in_target=merged_in_target,
        merged_check_error=merged_check_error,
        has_conflict=has_conflict,
        conflict_check_method=conflict_check_method,
        conflict_check_error=conflict_check_error,
    )

    config_dict = _build_backport_runtime_config(
        item_config=item_config,
        base_config=base_config,
        base_project_dir=base_project_dir,
        base_target_path=base_target_path,
        fixed_commit=fixed_commit,
        target_branch=target_branch,
        tag=tag,
        commit_id=commit_id,
        args=args,
    )

    did_backport = False
    try:
        patch_path = item_config.get("patch_path") or item_config.get("original_patch_path") or patch_path
        backport_result, did_backport, refreshed_status = _execute_backport_batch_action(
            is_report_config=is_report_config,
            is_merge_commit=is_merge_commit,
            merged_in_target=merged_in_target,
            has_conflict=has_conflict,
            conflict_check_method=conflict_check_method,
            conflict_check_error=conflict_check_error,
            item_config=item_config,
            tag=tag,
            commit_id=commit_id,
            target_branch=target_branch,
            fixed_commit=fixed_commit,
            patch_path=patch_path,
            config_dict=config_dict,
            base_target_path=base_target_path,
            base_project_dir=base_project_dir,
            args=args,
        )
        if refreshed_status:
            merged_in_target = refreshed_status.get("merged_in_target")
            merged_check_error = refreshed_status.get("merged_check_error")
            has_conflict = refreshed_status.get("has_conflict")
            conflict_check_method = refreshed_status.get("conflict_check_method")
            conflict_check_error = refreshed_status.get("conflict_check_error")
        # report 配置下，直接应用成功代表该提交已在目标分支生效，状态应及时回写
        if (
            is_report_config
            and backport_result.get("status") == "success"
            and not did_backport
            and not getattr(args, "stop_at_first_conflict", False)
        ):
            merged_in_target = True
            merged_check_error = None
            has_conflict = False
            if not conflict_check_method:
                conflict_check_method = "apply"
            conflict_check_error = None

        result = None
        if did_backport:
            result = _build_batch_summary_result(
                tag=tag,
                commit_id=commit_id,
                target_branch=target_branch,
                fixed_commit=fixed_commit,
                merged_in_target=merged_in_target,
                merged_check_error=merged_check_error,
                has_conflict=has_conflict,
                conflict_check_method=conflict_check_method,
                conflict_check_error=conflict_check_error,
                backport_result=backport_result,
            )
        report_item = _build_backport_batch_report_item(
            commit_id=commit_id,
            input_commit=input_commit,
            commit_title=commit_title,
            committed_datetime=committed_datetime,
            git_describe=git_describe,
            target_branch=target_branch,
            backport_result=backport_result,
            item_config=item_config,
            patch_path=patch_path,
            merged_in_target=merged_in_target,
            merged_check_error=merged_check_error,
            is_merge_commit=is_merge_commit,
            has_conflict=has_conflict,
            conflict_check_method=conflict_check_method,
            conflict_check_error=conflict_check_error,
            base_config=base_config,
            args=args,
        )
        return {
            "skip": False,
            "result": result,
            "report_item": report_item,
            "did_backport": did_backport,
        }
    except Exception as e:
        result = _build_backport_batch_failed_result(
            did_backport=did_backport,
            tag=tag,
            commit_id=commit_id,
            target_branch=target_branch,
            fixed_commit=fixed_commit,
            has_conflict=has_conflict,
            conflict_check_method=conflict_check_method,
            conflict_check_error=conflict_check_error,
            error=e,
        )
        report_item = {
            "commit": commit_id,
            "input_commit": input_commit,
            "commit_title": commit_title,
            "committed_datetime": committed_datetime,
            "git_describe": git_describe,
            "target_branch": target_branch,
            "status": "failed",
            "merged_in_target": merged_in_target,
            "merged_check_error": merged_check_error,
            "is_merge_commit": is_merge_commit,
            "has_conflict": has_conflict,
            "conflict_check_method": conflict_check_method,
            "conflict_check_error": conflict_check_error,
            "original_patch_path": None,
            "backported_patch_path": None,
            "patch_path": None,
            "error": str(e),
        }
        return {
            "skip": False,
            "result": result,
            "report_item": report_item,
            "did_backport": did_backport,
        }


def _build_backport_batch_report(
    base_config,
    base_project_dir,
    base_target_path,
    default_target_branch,
    llm_provider,
    llm_base_url,
    llm_model_name,
    backport_engine,
    report_items,
):
    # 计算有效值
    effective_provider = base_config.get("llm_provider") or llm_provider
    effective_base_url = base_config.get("llm_base_url") or llm_base_url
    effective_model_name = base_config.get("llm_model_name") or llm_model_name

    report = {
        "project": base_config.get("project", "linux"),
        "project_url": base_config.get("project_url", ""),
        "project_dir": base_project_dir,
        "target_path": base_target_path,
        "target_release": (
            base_config.get("target_branch")
            or base_config.get("target_release")
            or default_target_branch
        ),
        "patch_dataset_dir": base_config.get("patch_dataset_dir"),
        "llm_provider": effective_provider,
        "backport_engine": backport_engine,
        "commit_message_template": base_config.get("commit_message_template") or DEFAULT_COMMIT_MESSAGE_TEMPLATE,
        "linux_repo_path": base_config.get("linux_repo_path") or DEFAULT_LINUX_REPO_PATH,
    }

    # 在 llm_provider 之后添加 llm_base_url 和 llm_model_name（如果有值）
    if effective_base_url:
        report["llm_base_url"] = effective_base_url
    if effective_model_name:
        report["llm_model_name"] = effective_model_name

    # commits 放在最后
    report["commits"] = report_items
    return report


def _write_backport_batch_report(config_path, is_report_config, report):
    report_path = config_path if is_report_config else config_path + ".report.yml"
    with open(report_path, "w", encoding="utf-8") as f:
        yaml.safe_dump(report, f, allow_unicode=True, sort_keys=False)


def _prepare_backport_batch_context(args):
    config, commit_items = _load_backport_batch_config(args.backport_config)
    is_report_config = _is_report_config(args.backport_config, commit_items)
    report_output_path = args.backport_config
    logger.info("[backport-batch] 配置类型: %s", "report" if is_report_config else "raw")

    if args.interactive:
        if not is_report_config:
            logger.info("[backport-batch] 交互模式仅支持 report 配置，已跳过")
        else:
            interactive_action = _interactive_adjust_merged_in_target(
                args.backport_config, config, commit_items
            )
            maybe_output_path = config.pop("_interactive_output_report_path", "")
            if maybe_output_path:
                report_output_path = maybe_output_path
            if interactive_action == "quit":
                return None
            if interactive_action == "save":
                logger.info("[backport-batch] 已保存筛选配置，按用户指令退出，不继续执行")
                return None

    base_config = {k: v for k, v in config.items() if k != "commits"}
    base_project_dir = base_config.get("project_dir")
    if not base_project_dir:
        raise ValueError("backport-batch 必须提供源代码仓路径 project_dir")
    base_target_path = base_config.get("target_path")
    if not base_target_path:
        raise ValueError("backport-batch 必须提供目标仓路径 target_path")

    sorted_items, sort_errors = _resolve_sorted_backport_items(
        commit_items,
        is_report_config,
        base_project_dir,
        base_config,
        args,
    )
    return BackportBatchContext(
        config=config,
        is_report_config=is_report_config,
        report_output_path=report_output_path,
        base_config=base_config,
        base_project_dir=base_project_dir,
        base_target_path=base_target_path,
        sorted_items=sorted_items,
        sort_errors=sort_errors,
    )


def _execute_backport_batch_items(
    sorted_items,
    sort_errors,
    is_report_config,
    base_config,
    base_project_dir,
    base_target_path,
    default_target_branch,
    args,
):
    results = []
    report_items = []
    stop_at_first_conflict = bool(getattr(args, "stop_at_first_conflict", False))
    start_index = 0

    if stop_at_first_conflict and is_report_config:
        pending_indexes = [
            idx for idx, item in enumerate(sorted_items)
            if _is_pending_backport_item(item)
        ]
        if not pending_indexes:
            logger.info("[backport-batch] report 中没有 pending 条目，保持原 report 顺序返回")
            return results, [_copy_existing_report_item(item) for item in sorted_items]
        start_index = pending_indexes[0]

    for _, item, error_message in sort_errors:
        _append_sort_error_report_item(report_items, item, error_message)

    for idx, item in enumerate(sorted_items):
        if stop_at_first_conflict and is_report_config and idx < start_index:
            report_items.append(_copy_existing_report_item(item))
            continue

        processed = _process_backport_batch_item(
            item=item,
            is_report_config=is_report_config,
            base_config=base_config,
            base_project_dir=base_project_dir,
            base_target_path=base_target_path,
            default_target_branch=default_target_branch,
            args=args,
        )
        if processed.get("skip"):
            continue
        if processed.get("result"):
            results.append(processed["result"])
        report_items.append(processed["report_item"])
        if (
            stop_at_first_conflict
            and processed.get("report_item", {}).get("has_conflict") is True
        ):
            logger.info("[backport-batch] 检测到第一条冲突，停止后续检查并标记为 pending")
            _append_remaining_pending_items(
                report_items=report_items,
                remaining_items=sorted_items[idx + 1 :],
                is_report_config=is_report_config,
                base_config=base_config,
                default_target_branch=default_target_branch,
            )
            break
        if processed.get("did_backport"):
            logger.info("[backport-batch] 已执行回移植，停止后续处理以便检查 report.yml")
            _append_remaining_pending_items(
                report_items=report_items,
                remaining_items=sorted_items[idx + 1 :],
                is_report_config=is_report_config,
                base_config=base_config,
                default_target_branch=default_target_branch,
            )
            break
    return results, report_items
