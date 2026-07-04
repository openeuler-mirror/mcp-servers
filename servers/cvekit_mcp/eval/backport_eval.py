#!/usr/bin/env python3
"""Evaluate cvekit Mystique backports against the manual commits in a target PR.

Each source commit is evaluated independently.  Before a case starts, the
target repository is reset to the manual state immediately before the next
relevant manual commit.  This prevents an earlier AI result from affecting a
later case.
"""

from __future__ import annotations

import argparse
import fcntl
import json
import logging
import os
import re
import shlex
import subprocess
import sys
import time
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
import yaml


SCRIPT_DIR = Path(__file__).resolve().parent
TEMP_BRANCH = "eval/backport-eval"
OUTPUT_DIR = Path.cwd()
LOG_ROOT = SCRIPT_DIR / "logs"
RECOVERY_FILE = SCRIPT_DIR / "backport_eval_recovery.json"
LOCK_FILE = SCRIPT_DIR / "backport_eval.lock"
EVAL_NAME = "backport_eval"

# Evaluation order is computed once from all unique source commits using
# cvekit's describe ordering. A case raw config always contains exactly one
# exact SHA, so sorting inside that case cannot change evaluation order.
#
# cvekit has no "skip sorting" mode. Using "time" for a one-SHA case avoids
# describe/title resolution scanning the complete source branch. Do not add
# commit_title to the raw item: cvekit would rebuild the full title index.
DISCOVERY_COMMIT_SORT = "describe"
SINGLE_CASE_COMMIT_SORT = "time"

SOURCE_REF_PATTERNS = (
    re.compile(r"(?im)^\s*commit[:\s]+([0-9a-f]{7,40})\s+(?:upstream|openeuler)\s*$"),
    re.compile(r"(?im)^\s*(?:upstream|openeuler)\s+commit[:\s]+([0-9a-f]{7,40})\s*$"),
    re.compile(r"(?im)^\s*\[\s*(?:upstream|openeuler)\s+commit\s+([0-9a-f]{7,40})\s*\]\s*$"),
    re.compile(
        r"(?im)^\s*(?:source|upstream|mainline|original|origin|linux|openeuler|原始|源代码仓)"
        r".*?commit(?:\s+(?:id|hash))?\s*[:：]?\s+([0-9a-f]{7,40})\b"
    ),
    re.compile(
        r"(?im)^\s*commit(?:\s+(?:id|hash))?\s*[:：]?\s+([0-9a-f]{7,40})\b"
        r".*?(?:source|upstream|mainline|original|origin|linux|openeuler|原始|源代码仓)"
    ),
    re.compile(r"(?im)^\s*commit[:\s]+([0-9a-f]{7,40})\s*$"),
    re.compile(r"(?im)^\s*\(?cherry[ -]?picked from commit\s+([0-9a-f]{7,40})\)?\s*$"),
)
PATCH_FROM_RE = re.compile(r"(?m)^From\s+([0-9a-f]{7,40})\s+")
TARGET_COMMIT_HEADER_ALIASES = {
    "targetcommithash",
    "targetcommit",
    "manualcommithash",
    "manualcommit",
}
SOURCE_COMMIT_HEADER_ALIASES = {
    "commithash",
    "commit",
    "commitid",
    "sourcecommithash",
    "sourcecommit",
    "sourcecommitid",
}
SAFE_NAME_RE = re.compile(r"[^A-Za-z0-9_.-]+")
# Excel cells have a hard 32,767-character limit. Split oversized text across
# continuation rows so large patches/log snippets can still be exported.
EXCEL_CELL_LIMIT = 32767
LOGGER = logging.getLogger("backport_eval")
_TARGET_TITLE_INDEX_CACHE: dict[str, dict[str, list[str]]] = {}


class EvalError(RuntimeError):
    pass


@dataclass
class SourceCommit:
    describe_index: int
    excel_rows: list[int]
    target_commits: list[str]
    source_commit: str
    source_title: str
    git_describe: str


@dataclass
class ManualPatchFile:
    index: int
    path: Path
    target_commit: str
    message: str
    source_refs: list[str]


@dataclass
class ManualCommit:
    index: int
    commit: str
    parent: str
    title: str
    message: str
    source_refs: list[str]


@dataclass
class EvalCase:
    source: SourceCommit
    manual_pr_index: int | None
    manual_commit: str
    case_baseline: str
    expected_behavior: str


@dataclass
class Config:
    eval_name: str
    source_repo: Path
    source_branch: str
    source_excel: Path | None
    preprocessed_source_excel: Path | None
    target_repo: Path
    manual_patch_dir: Path | None
    manual_input_mode: str
    pr_url: str
    first_pr_commit: str
    last_pr_commit: str
    pr_baseline: str
    temp_branch: str
    cvekit: Path
    cvekit_workdir: Path
    output_dir: Path
    log_root: Path
    recovery_file: Path
    lock_file: Path
    api_key: str
    llm_provider: str
    llm_base_url: str
    llm_model_name: str
    signer_name: str
    signer_email: str
    case_limit: int


def setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def attach_run_log(log_path: Path) -> None:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    handler = logging.FileHandler(log_path, encoding="utf-8")
    handler.setFormatter(
        logging.Formatter("%(asctime)s %(levelname)s %(message)s", datefmt="%Y-%m-%d %H:%M:%S")
    )
    logging.getLogger().addHandler(handler)
    LOGGER.info("persistent run log: %s", log_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--discover-only",
        action="store_true",
        help="仅做输入检查、提交排序、映射关系和每个 case 基线推导，不执行实际评测",
    )
    parser.add_argument(
        "--case-limit",
        type=int,
        default=0,
        help="只运行前 N 个已独立准备好的 case；默认 0 表示运行全部",
    )
    parser.add_argument(
        "--resume",
        default="",
        help="从已有 run 目录断点续跑；已成功 checkpoint 的 case 会自动跳过",
    )
    parser.add_argument("--eval-name", default=EVAL_NAME)
    parser.add_argument("--source-repo", default="")
    parser.add_argument("--source-branch", default="")
    parser.add_argument(
        "--source-excel",
        default="",
        help="描述 source commit 的 Excel；不传时从 PR commit message 中提取源 commit",
    )
    parser.add_argument("--target-repo", default="")
    parser.add_argument(
        "--manual-patch-dir",
        default="",
        help="人工 backport patch 目录；传入后从 format-patch 文件读取人工提交序列，而不是从 PR commit range 读取",
    )
    parser.add_argument("--pr-url", default="")
    parser.add_argument("--first-pr-commit", default="", help="PR 模式必填；manual patch 目录模式可省略")
    parser.add_argument("--last-pr-commit", default="", help="PR 模式必填；manual patch 目录模式可省略")
    parser.add_argument("--temp-branch", default=TEMP_BRANCH)
    parser.add_argument("--cvekit", default="")
    parser.add_argument("--cvekit-workdir", default="")
    parser.add_argument("--output-dir", default=str(OUTPUT_DIR))
    parser.add_argument("--log-root", default=str(LOG_ROOT))
    parser.add_argument("--recovery-file", default="")
    parser.add_argument("--lock-file", default="")
    parser.add_argument(
        "--api-key",
        default=os.environ.get("API_KEY") or os.environ.get("OPENAI_KEY", ""),
        help="LLM API key passed through to cvekit; defaults to API_KEY/OPENAI_KEY",
    )
    parser.add_argument("--llm-provider", default=os.environ.get("LLM_PROVIDER", ""))
    parser.add_argument("--llm-base-url", default=os.environ.get("LLM_BASE_URL", ""))
    parser.add_argument(
        "--llm-model-name",
        default=os.environ.get("LLM_MODEL_NAME") or os.environ.get("MODEL_NAME", ""),
    )
    parser.add_argument("--signer-name", default=os.environ.get("SIGNER_NAME", "AI Eval"))
    parser.add_argument(
        "--signer-email",
        default=os.environ.get("SIGNER_EMAIL", "ai-eval@example.invalid"),
    )
    return parser.parse_args()


def build_config(args: argparse.Namespace) -> Config:
    required = {
        "source_repo": args.source_repo,
        "source_branch": args.source_branch,
        "target_repo": args.target_repo,
        "cvekit": args.cvekit,
    }
    manual_patch_dir_arg = args.manual_patch_dir.strip()
    has_pr_range = bool(args.first_pr_commit.strip() and args.last_pr_commit.strip())
    has_partial_pr_range = bool(args.first_pr_commit.strip()) != bool(args.last_pr_commit.strip())
    if has_partial_pr_range:
        raise EvalError("--first-pr-commit and --last-pr-commit must be passed together")
    if manual_patch_dir_arg:
        required["manual_patch_dir"] = manual_patch_dir_arg
    elif has_pr_range:
        required["pr_url"] = args.pr_url
    else:
        required["source_excel"] = args.source_excel
    missing = [name for name, value in required.items() if not str(value).strip()]
    if missing:
        raise EvalError(
            "missing required arguments: " + ", ".join(f"--{name.replace('_', '-')}" for name in missing)
        )
    eval_name = safe_name(args.eval_name) or EVAL_NAME
    source_repo = Path(args.source_repo).expanduser().resolve()
    target_repo = Path(args.target_repo).expanduser().resolve()
    manual_patch_dir = (
        Path(manual_patch_dir_arg).expanduser().resolve()
        if manual_patch_dir_arg
        else None
    )
    source_excel = (
        Path(args.source_excel).expanduser().resolve()
        if args.source_excel.strip()
        else None
    )
    cvekit = Path(args.cvekit).expanduser().resolve()
    cvekit_workdir = (
        Path(args.cvekit_workdir).expanduser().resolve()
        if args.cvekit_workdir.strip()
        else cvekit.parent
    )
    output_dir = Path(args.output_dir).expanduser().resolve()
    log_root = Path(args.log_root).expanduser().resolve()
    recovery_file = (
        Path(args.recovery_file).expanduser().resolve()
        if args.recovery_file
        else RECOVERY_FILE.with_name(f"{eval_name}_recovery.json")
    )
    lock_file = (
        Path(args.lock_file).expanduser().resolve()
        if args.lock_file
        else LOCK_FILE.with_name(f"{eval_name}.lock")
    )
    if manual_patch_dir is not None:
        manual_input_mode = "manual_patch_dir"
        manual_patches = read_manual_patch_files(target_repo, source_repo, manual_patch_dir)
        if not manual_patches:
            raise EvalError(f"no manual patch files found in {manual_patch_dir}")
        first_pr_commit = manual_patches[0].target_commit
        last_pr_commit = manual_patches[-1].target_commit
        pr_url = args.pr_url.strip() or f"manual-patch-dir:{manual_patch_dir}"
        pr_baseline = git(target_repo, "rev-parse", f"{first_pr_commit}^")
    elif has_pr_range:
        manual_input_mode = "pr_range"
        first_pr_commit = git_verify(target_repo, args.first_pr_commit.strip())
        last_pr_commit = git_verify(target_repo, args.last_pr_commit.strip())
        pr_url = args.pr_url.strip()
        pr_baseline = git(target_repo, "rev-parse", f"{first_pr_commit}^")
    else:
        if source_excel is None:
            raise EvalError(
                "missing manual input: pass --manual-patch-dir, pass "
                "--first-pr-commit/--last-pr-commit, or pass --source-excel "
                "with a target commit hash column"
            )
        has_target_commit_col = find_excel_header_column(source_excel, TARGET_COMMIT_HEADER_ALIASES) is not None
        manual_input_mode = "excel_target_commits" if has_target_commit_col else "excel_title_lookup"
        target_commits = (
            read_excel_target_commits(target_repo, source_excel)
            if has_target_commit_col
            else read_excel_target_commits_by_title(target_repo, source_excel)
        )
        if not target_commits:
            raise EvalError(
                "source Excel did not resolve to any target commits; pass --manual-patch-dir "
                "or --first-pr-commit/--last-pr-commit, add a target commit hash column, "
                "or make sure commit titles exist on target HEAD"
            )
        first_pr_commit = target_commits[0]
        last_pr_commit = target_commits[-1]
        pr_url = args.pr_url.strip() or f"{manual_input_mode}:{source_excel}"
        pr_baseline = git(target_repo, "rev-parse", f"{first_pr_commit}^")
    return Config(
        eval_name=eval_name,
        source_repo=source_repo,
        source_branch=args.source_branch.strip(),
        source_excel=source_excel,
        preprocessed_source_excel=None,
        target_repo=target_repo,
        manual_patch_dir=manual_patch_dir,
        manual_input_mode=manual_input_mode,
        pr_url=pr_url,
        first_pr_commit=first_pr_commit,
        last_pr_commit=last_pr_commit,
        pr_baseline=pr_baseline,
        temp_branch=args.temp_branch.strip(),
        cvekit=cvekit,
        cvekit_workdir=cvekit_workdir,
        output_dir=output_dir,
        log_root=log_root,
        recovery_file=recovery_file,
        lock_file=lock_file,
        api_key=args.api_key.strip(),
        llm_provider=args.llm_provider.strip(),
        llm_base_url=args.llm_base_url.strip(),
        llm_model_name=args.llm_model_name.strip(),
        signer_name=args.signer_name.strip(),
        signer_email=args.signer_email.strip(),
        case_limit=max(0, args.case_limit),
    )


def acquire_run_lock(config: Config) -> Any:
    config.lock_file.parent.mkdir(parents=True, exist_ok=True)
    handle = config.lock_file.open("a+", encoding="utf-8")
    try:
        fcntl.flock(handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
    except BlockingIOError as exc:
        handle.seek(0)
        owner = handle.read().strip() or "unknown"
        handle.close()
        raise EvalError(f"another {config.eval_name} evaluation is running: {owner}") from exc
    handle.seek(0)
    handle.truncate()
    handle.write(f"pid={os.getpid()}\n")
    handle.flush()
    return handle


def run(
    cmd: list[str],
    *,
    cwd: Path | None = None,
    check: bool = True,
    input_text: str | None = None,
    log_path: Path | None = None,
) -> subprocess.CompletedProcess[str]:
    def redact_cmd(parts: list[str]) -> str:
        redacted: list[str] = []
        hide_next = False
        for part in parts:
            if hide_next:
                redacted.append("***")
                hide_next = False
                continue
            if part == "--api-key":
                redacted.append(part)
                hide_next = True
            elif part.startswith("--api-key="):
                redacted.append("--api-key=***")
            else:
                redacted.append(part)
        return " ".join(shlex.quote(part) for part in redacted)

    safe_cmd = redact_cmd(cmd)
    LOGGER.debug("run: %s", safe_cmd)
    started = time.monotonic()
    if log_path:
        LOGGER.info("COMMAND START: output=%s", log_path)
    completed = subprocess.run(
        cmd,
        cwd=str(cwd) if cwd else None,
        input=input_text,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        env=os.environ.copy(),
    )
    if log_path:
        log_path.parent.mkdir(parents=True, exist_ok=True)
        log_path.write_text(
            f"$ {safe_cmd}\n\n"
            f"--- stdout ---\n{completed.stdout}\n"
            f"--- stderr ---\n{completed.stderr}\n",
            encoding="utf-8",
        )
        LOGGER.info(
            "COMMAND FINISH: output=%s exit_code=%d elapsed=%.1fs",
            log_path,
            completed.returncode,
            time.monotonic() - started,
        )
    if check and completed.returncode != 0:
        raise EvalError(
            f"command failed ({completed.returncode}): "
            f"{safe_cmd}\n"
            f"stdout:\n{completed.stdout}\nstderr:\n{completed.stderr}"
        )
    return completed


def git(repo: Path, *args: str, check: bool = True, input_text: str | None = None) -> str:
    return run(
        ["git", "-C", str(repo), *args],
        check=check,
        input_text=input_text,
    ).stdout.strip()


def git_verify(repo: Path, revision: str) -> str:
    return git(repo, "rev-parse", "--verify", f"{revision}^{{commit}}")


def normalize_title(title: str) -> str:
    return " ".join(title.strip().split())


def safe_name(value: str) -> str:
    return SAFE_NAME_RE.sub("_", value).strip("_")[:80] or "case"


def extract_source_refs(message: str) -> list[str]:
    refs: list[str] = []
    seen: set[str] = set()
    for pattern in SOURCE_REF_PATTERNS:
        for match in pattern.finditer(message):
            commit = match.group(1)
            if commit not in seen:
                seen.add(commit)
                refs.append(commit)
    return refs


def patch_file_sort_key(path: Path) -> tuple[int, str]:
    match = re.match(r"^(\d+)", path.name)
    return (int(match.group(1)) if match else 10**9, path.name)


def read_manual_patch_files(
    target_repo: Path,
    source_repo: Path,
    manual_patch_dir: Path,
) -> list[ManualPatchFile]:
    if not manual_patch_dir.is_dir():
        raise EvalError(f"manual patch dir is not a directory: {manual_patch_dir}")
    # 按 patch 文件名前缀数字排序，比如 0001-...patch, 0002-...patch 
    patch_paths = sorted(manual_patch_dir.glob("*.patch"), key=patch_file_sort_key)
    patches: list[ManualPatchFile] = []
    for index, path in enumerate(patch_paths, start=1):
        text = path.read_text(encoding="utf-8", errors="replace")
        match = PATCH_FROM_RE.search(text)
        if not match:
            raise EvalError(f"manual patch has no format-patch From commit line: {path}")
        target_commit = git_verify(target_repo, match.group(1))
        refs: list[str] = []
        for ref in extract_source_refs(text):
            try:
                refs.append(git_verify(source_repo, ref))
            except EvalError:
                LOGGER.warning(
                    "manual patch %s references source hash not present locally: %s",
                    path.name,
                    ref,
                )
        patches.append(
            ManualPatchFile(
                index=index,
                path=path,
                target_commit=target_commit,
                message=text,
                source_refs=refs,
            )
        )
    return patches


def normalize_commit_title(title: str) -> str:
    """Normalize downstream/source title variants for exact subject matching.

    Downstream kernel repos often prefix titles with tags like
    ``downstream:``, ``openeuler:`` or ``OLK:`` that are not present in the
    upstream source repo.  Strip those markers and collapse whitespace before
    comparing target and source subjects.
    """
    normalized = normalize_title(title)
    lower = normalized.lower()
    known_prefixes = ("downstream:", "openeuler:", "olk:", "!this_is_not_the_commit:")
    for prefix in known_prefixes:
        if lower.startswith(prefix):
            return normalized[len(prefix):].lstrip()
    return normalized


def is_revert_subject(title: str) -> bool:
    return normalize_title(title).lower().startswith("revert ")


def find_source_commits_by_normalized_title(
    config: Config,
    target_title: str,
) -> list[str]:
    target_title_norm = normalize_commit_title(target_title)
    grep_title = target_title_norm
    found = git(
        config.source_repo,
        "log",
        f"{config.source_branch}",
        "--format=%H%x00%s",
        "--no-merges",
        "--fixed-strings",
        f"--grep={grep_title}",
        check=False,
    ).splitlines()
    commits: list[str] = []
    for line in found:
        found_commit, _, subject = line.partition("\x00")
        if is_revert_subject(subject):
            continue
        # 只包含标题完全一样的
        if normalize_commit_title(subject) != target_title_norm:
            continue
        try:
            full = git_verify(config.source_repo, found_commit)
        except EvalError:
            LOGGER.warning(
                "fallback found commit %s but cannot verify in source repo",
                found_commit[:12],
            )
            continue
        if full not in commits:
            commits.append(full)
    return commits


def source_refs_from_target_commit(
    config: Config,
    target_commit: str,
    manual_by_commit: dict[str, ManualCommit],
) -> list[str]:
    full_target_commit = git_verify(config.target_repo, target_commit)
    manual = manual_by_commit.get(full_target_commit)
    if manual and manual.source_refs:
        return manual.source_refs
    message = git(config.target_repo, "show", "-s", "--format=%B", full_target_commit)
    refs: list[str] = []
    # target commit -> source commit: 1. commit message里查找upstream sha 
    for ref in extract_source_refs(message):
        try:
            refs.append(git_verify(config.source_repo, ref))
        except EvalError:
            LOGGER.warning(
                "target commit %s references source hash not present locally: %s",
                full_target_commit[:12],
                ref,
            )
    if not refs:
        # 匹配不到的话再在源仓库历史里面搜标题
        title = git(config.target_repo, "show", "-s", "--format=%s", full_target_commit)
        LOGGER.info(
            "fallback: searching source repo by normalized title for target commit %s title=%s",
            full_target_commit[:12],
            normalize_commit_title(title),
        )
        matches = find_source_commits_by_normalized_title(config, title)
        if matches:
            # git log returns newest matches first for the selected source branch.
            refs.append(matches[0])
            if len(matches) == 1:
                LOGGER.info(
                    "fallback found unique source commit %s for target commit %s",
                    matches[0][:12],
                    full_target_commit[:12],
                )
            else:
                LOGGER.warning(
                    "fallback title search found multiple source commits for target commit %s; "
                    "using first=%s candidates=%s",
                    full_target_commit[:12],
                    matches[0][:12],
                    ", ".join(commit[:12] for commit in matches),
                )
    return refs


def sort_source_commits(
    config: Config,
    by_commit: dict[str, dict[str, Any]],
) -> list[SourceCommit]:
    try:
        sys.path.insert(0, str(config.cvekit_workdir))
        from cvekit.utils.backport_sort import sort_commit_items

        sorted_items, errors = sort_commit_items(
            [{"commit": commit} for commit in by_commit],
            str(config.source_repo),
            config.source_branch,
            DISCOVERY_COMMIT_SORT,
        )
    except Exception as exc:
        raise EvalError(f"failed to sort source commits with cvekit describe rules: {exc}") from exc
    if errors:
        raise EvalError(f"cvekit describe sorting failed: {errors}")

    sources: list[SourceCommit] = []
    for index, item in enumerate(sorted_items, start=1):
        commit = str(item["commit"])
        entry = by_commit[commit]
        sources.append(
            SourceCommit(
                describe_index=index,
                excel_rows=entry["excel_rows"],
                target_commits=entry.get("target_commits", []),
                source_commit=commit,
                source_title=entry["source_title"],
                git_describe=str(item.get("git_describe") or ""),
            )
        )
    return sources


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def find_header_column(headers: list[Any], aliases: set[str]) -> int | None:
    normalized = [normalize_header(value) for value in headers]
    for index, header in enumerate(normalized):
        if header in aliases:
            return index
    return None


def repo_key(repo: Path) -> str:
    return str(Path(repo).expanduser().resolve())


def find_excel_header_column(source_excel: Path, aliases: set[str]) -> int | None:
    workbook = openpyxl.load_workbook(source_excel, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    return find_header_column([cell.value for cell in sheet[1]], aliases)


def read_excel_target_commits(target_repo: Path, source_excel: Path) -> list[str]:
    """Read manual backport commits directly from an Excel target commit column.

    In ``excel_target_commits`` mode, the Excel file is the authoritative
    evaluation sample list.  Each non-empty ``target commit hash`` cell is a
    manual backport commit; its parent becomes the case baseline and
    ``git diff <target>^ <target>`` remains the manual patch used for later
    comparison.  No PR range or exported format-patch directory is needed.
    """
    workbook = openpyxl.load_workbook(source_excel, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header_values = [cell.value for cell in sheet[1]]
    target_commit_col = find_header_column(header_values, TARGET_COMMIT_HEADER_ALIASES)
    if target_commit_col is None:
        raise EvalError(
            "source Excel has no target commit hash column; pass --manual-patch-dir "
            "or --first-pr-commit/--last-pr-commit, or add a target commit hash column"
        )

    commits: list[str] = []
    seen: set[str] = set()
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        commit_input = str(values[target_commit_col] or "").strip()
        if not commit_input:
            continue
        full_commit = git_verify(target_repo, commit_input)
        if full_commit not in seen:
            seen.add(full_commit)
            commits.append(full_commit)
        else:
            LOGGER.warning(
                "Excel row %d repeats target commit %s; keeping the first occurrence",
                row_number,
                full_commit[:12],
            )
    return commits


def read_excel_source_rows(source_excel: Path) -> list[dict[str, str | int]]:
    workbook = openpyxl.load_workbook(source_excel, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header_values = [cell.value for cell in sheet[1]]
    title_col = find_header_column(header_values, {"committitle", "title", "subject"})
    commit_col = find_header_column(header_values, SOURCE_COMMIT_HEADER_ALIASES)
    if title_col is None:
        title_col = 0

    rows: list[dict[str, str | int]] = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        title = normalize_title(str(values[title_col] or ""))
        commit_input = (
            str(values[commit_col] or "").strip()
            if commit_col is not None and commit_col < len(values)
            else ""
        )
        if not title and not commit_input:
            continue
        rows.append({"row_number": row_number, "title": title, "commit": commit_input})
    return rows


def target_title_index(target_repo: Path) -> dict[str, list[str]]:
    key = repo_key(target_repo)
    if key in _TARGET_TITLE_INDEX_CACHE:
        return _TARGET_TITLE_INDEX_CACHE[key]

    LOGGER.info("building target title index from HEAD: %s", target_repo)
    completed = subprocess.run(
        ["git", "-C", str(target_repo), "log", "HEAD", "--format=%H%x00%s"],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
    )
    if completed.returncode != 0:
        raise EvalError(
            "git log failed while building target title index: "
            + completed.stderr.decode("utf-8", errors="replace").strip()
        )
    output = completed.stdout.decode("utf-8", errors="replace")
    index: dict[str, list[str]] = {}
    for line in output.splitlines():
        commit, _, subject = line.partition("\x00")
        if not commit:
            continue
        title_norm = normalize_commit_title(subject)
        if not title_norm:
            continue
        commits = index.setdefault(title_norm, [])
        if commit not in commits:
            commits.append(commit)
    _TARGET_TITLE_INDEX_CACHE[key] = index
    LOGGER.info("target title index built: titles=%d repo=%s", len(index), target_repo)
    return index


def find_target_commit_by_normalized_title(target_repo: Path, title: str) -> str:
    title_norm = normalize_commit_title(title)
    matches = target_title_index(target_repo).get(title_norm, [])
    if not matches:
        raise EvalError(f"target branch HEAD has no commit matching Excel title: {title}")
    if len(matches) > 1:
        LOGGER.warning(
            "target title matched multiple commits; using newest title=%s candidates=%s",
            title_norm,
            ", ".join(commit[:12] for commit in matches),
        )
    return matches[0]


def read_excel_target_commits_by_title(target_repo: Path, source_excel: Path) -> list[str]:
    commits: list[str] = []
    seen: set[str] = set()
    for row in read_excel_source_rows(source_excel):
        title = str(row["title"])
        if not title:
            raise EvalError(f"Excel row {row['row_number']} has no commit title for target lookup")
        commit = find_target_commit_by_normalized_title(target_repo, title)
        if commit not in seen:
            seen.add(commit)
            commits.append(commit)
        else:
            LOGGER.warning(
                "Excel row %s maps to repeated target commit %s; keeping the first occurrence",
                row["row_number"],
                commit[:12],
            )
    return commits


def read_source_commits_from_excel(
    config: Config,
    manuals: list[ManualCommit],
    preprocessed_excel_path: Path | None = None,
) -> tuple[list[SourceCommit], int, str]:
    if config.source_excel is None:
        raise EvalError("source Excel is not configured")
    workbook = openpyxl.load_workbook(config.source_excel)
    sheet = workbook[workbook.sheetnames[0]]
    header_values = [cell.value for cell in sheet[1]]
    title_col = find_header_column(header_values, {"committitle", "title", "subject"})
    commit_col = find_header_column(header_values, SOURCE_COMMIT_HEADER_ALIASES)
    target_commit_col = find_header_column(
        header_values,
        TARGET_COMMIT_HEADER_ALIASES,
    )
    if title_col is None:
        title_col = 0
    if commit_col is None:
        commit_col = len(header_values)
        sheet.cell(row=1, column=commit_col + 1, value="commit hash")

    manual_by_commit = {manual.commit: manual for manual in manuals}
    by_commit: dict[str, dict[str, Any]] = {}
    input_count = 0
    derived_count = 0
    skipped_count = 0
    for row_number in range(2, sheet.max_row + 1):
        title = normalize_title(str(sheet.cell(row=row_number, column=title_col + 1).value or ""))
        commit_input = str(sheet.cell(row=row_number, column=commit_col + 1).value or "").strip()
        target_commit_input = (
            str(sheet.cell(row=row_number, column=target_commit_col + 1).value or "").strip()
            if target_commit_col is not None
            else ""
        )
        if not title and not commit_input and not target_commit_input:
            continue
        if not commit_input:
            if not target_commit_input:
                raise EvalError(
                    f"Excel row {row_number} has no commit hash and no target commit hash"
                )
            full_target_commit = git_verify(config.target_repo, target_commit_input)
            source_refs = source_refs_from_target_commit(config, full_target_commit, manual_by_commit)
            if not source_refs:
                skipped_count += 1
                LOGGER.warning(
                    "Excel row %d target commit %s has no source commit reference; skipping",
                    row_number,
                    target_commit_input,
                )
                continue
            if len(source_refs) > 1:
                raise EvalError(
                    f"Excel row {row_number} target commit {target_commit_input} "
                    f"maps to multiple source commits: {source_refs}"
                )
            commit_input = source_refs[0]
            sheet.cell(row=row_number, column=commit_col + 1, value=commit_input)
            derived_count += 1
        full_target_commit = (
            git_verify(config.target_repo, target_commit_input)
            if target_commit_input
            else ""
        )
        full_commit = git_verify(config.source_repo, commit_input)
        input_count += 1
        actual_title = normalize_title(git(config.source_repo, "show", "-s", "--format=%s", full_commit))
        entry = by_commit.setdefault(
            full_commit,
            {"excel_rows": [], "target_commits": [], "source_title": actual_title or title},
        )
        entry["excel_rows"].append(row_number)
        if full_target_commit and full_target_commit not in entry["target_commits"]:
            entry["target_commits"].append(full_target_commit)
    if derived_count and preprocessed_excel_path is not None:
        preprocessed_excel_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(preprocessed_excel_path)
        config.preprocessed_source_excel = preprocessed_excel_path
        LOGGER.info(
            "preprocessed source Excel written: %s; derived commit hash rows=%d skipped rows=%d",
            preprocessed_excel_path,
            derived_count,
            skipped_count,
        )
    return sort_source_commits(config, by_commit), input_count, "excel"


def read_source_commits_from_pr_messages(
    config: Config,
    manuals: list[ManualCommit],
) -> tuple[list[SourceCommit], int, str]:
    by_commit: dict[str, dict[str, Any]] = {}
    input_count = 0
    for manual in manuals:
        for source_commit in manual.source_refs:
            input_count += 1
            actual_title = normalize_title(
                git(config.source_repo, "show", "-s", "--format=%s", source_commit)
            )
            entry = by_commit.setdefault(
                source_commit,
                {"excel_rows": [], "source_title": actual_title},
            )
            entry["excel_rows"].append(manual.index)
    if not by_commit:
        raise EvalError(
            "no source commits found in PR commit messages; pass --source-excel "
            "or add source commit references to the PR commits"
        )
    return sort_source_commits(config, by_commit), input_count, "pr_messages"


def read_manual_commits(config: Config) -> list[ManualCommit]:
    if config.manual_input_mode == "excel_target_commits":
        return read_manual_commits_from_excel_target_commits(config)
    if config.manual_input_mode == "excel_title_lookup":
        return read_manual_commits_from_excel_title_lookup(config)
    if config.manual_patch_dir is not None:
        return read_manual_commits_from_patch_dir(config)

    chain = git(
        config.target_repo,
        "rev-list",
        "--reverse",
        f"{config.pr_baseline}..{config.last_pr_commit}",
    ).splitlines()
    if not chain:
        raise EvalError("PR commit chain is empty")
    first_parent = git(config.target_repo, "rev-parse", f"{chain[0]}^")
    if first_parent != config.pr_baseline:
        raise EvalError(
            f"first manual commit parent is {first_parent}, expected {config.pr_baseline}"
        )
    manuals: list[ManualCommit] = []
    for index, commit in enumerate(chain, start=1):
        message = git(config.target_repo, "show", "-s", "--format=%B", commit)
        refs: list[str] = []
        for ref in extract_source_refs(message):
            try:
                refs.append(git_verify(config.source_repo, ref))
            except EvalError:
                LOGGER.warning(
                    "manual commit %s references source hash not present locally: %s",
                    commit[:12],
                    ref,
                )
        manuals.append(
            ManualCommit(
                index=index,
                commit=commit,
                parent=git(config.target_repo, "rev-parse", f"{commit}^"),
                title=normalize_title(git(config.target_repo, "show", "-s", "--format=%s", commit)),
                message=message,
                source_refs=refs,
            )
        )
    return manuals


def read_manual_commits_from_excel_title_lookup(config: Config) -> list[ManualCommit]:
    if config.source_excel is None:
        raise EvalError("source Excel is required for excel_title_lookup mode")

    manuals: list[ManualCommit] = []
    seen_targets: set[str] = set()
    for row in read_excel_source_rows(config.source_excel):
        title = str(row["title"])
        commit_input = str(row["commit"])
        if not title:
            raise EvalError(f"Excel row {row['row_number']} has no commit title for target lookup")
        target_commit = find_target_commit_by_normalized_title(config.target_repo, title)
        if target_commit in seen_targets:
            continue
        seen_targets.add(target_commit)

        refs: list[str] = []
        if commit_input:
            refs.append(git_verify(config.source_repo, commit_input))
        message = git(config.target_repo, "show", "-s", "--format=%B", target_commit)
        if not refs:
            for ref in extract_source_refs(message):
                try:
                    refs.append(git_verify(config.source_repo, ref))
                except EvalError:
                    LOGGER.warning(
                        "manual commit %s references source hash not present locally: %s",
                        target_commit[:12],
                        ref,
                    )
        manuals.append(
            ManualCommit(
                index=len(manuals) + 1,
                commit=target_commit,
                parent=git(config.target_repo, "rev-parse", f"{target_commit}^"),
                title=normalize_title(git(config.target_repo, "show", "-s", "--format=%s", target_commit)),
                message=message,
                source_refs=refs,
            )
        )
    return manuals


def read_manual_commits_from_excel_target_commits(config: Config) -> list[ManualCommit]:
    """Build manual commits from Excel target commit hash cells.

    This mode is for workbooks where every row already names the manual
    backport commit.  The exported patch directory is intentionally bypassed:
    the target repository is the source of truth for commit parent, subject,
    message, and later manual patch diff generation.
    """
    if config.source_excel is None:
        raise EvalError("source Excel is required for excel_target_commits mode")
    target_commits = read_excel_target_commits(config.target_repo, config.source_excel)
    manuals: list[ManualCommit] = []
    for index, commit in enumerate(target_commits, start=1):
        message = git(config.target_repo, "show", "-s", "--format=%B", commit)
        refs: list[str] = []
        for ref in extract_source_refs(message):
            try:
                refs.append(git_verify(config.source_repo, ref))
            except EvalError:
                LOGGER.warning(
                    "manual commit %s references source hash not present locally: %s",
                    commit[:12],
                    ref,
                )
        manuals.append(
            ManualCommit(
                index=index,
                commit=commit,
                parent=git(config.target_repo, "rev-parse", f"{commit}^"),
                title=normalize_title(git(config.target_repo, "show", "-s", "--format=%s", commit)),
                message=message,
                source_refs=refs,
            )
        )
    return manuals


def read_manual_commits_from_patch_dir(config: Config) -> list[ManualCommit]:
    if config.manual_patch_dir is None:
        raise EvalError("manual patch dir is not configured")
    patch_files = read_manual_patch_files(
        config.target_repo,
        config.source_repo,
        config.manual_patch_dir,
    )
    if not patch_files:
        raise EvalError(f"manual patch dir is empty: {config.manual_patch_dir}")

    manuals: list[ManualCommit] = []
    previous_commit = config.pr_baseline
    for patch in patch_files:
        parent = git(config.target_repo, "rev-parse", f"{patch.target_commit}^")
        if parent != previous_commit:
            LOGGER.warning(
                "manual patch commits are not a linear chain in target repo: "
                "%s parent is %s, previous patch commit is %s; using this patch parent as case baseline",
                patch.path.name,
                parent,
                previous_commit,
            )
        title = normalize_title(git(config.target_repo, "show", "-s", "--format=%s", patch.target_commit))
        message = git(config.target_repo, "show", "-s", "--format=%B", patch.target_commit)
        refs = list(patch.source_refs)
        if not refs:
            for ref in extract_source_refs(message):
                try:
                    refs.append(git_verify(config.source_repo, ref))
                except EvalError:
                    LOGGER.warning(
                        "manual commit %s references source hash not present locally: %s",
                        patch.target_commit[:12],
                        ref,
                    )
        manuals.append(
            ManualCommit(
                index=patch.index,
                commit=patch.target_commit,
                parent=parent,
                title=title,
                message=patch.message or message,
                source_refs=refs,
            )
        )
        previous_commit = patch.target_commit
    if manuals[-1].commit != config.last_pr_commit:
        raise EvalError(
            f"last manual patch commit is {manuals[-1].commit}, expected {config.last_pr_commit}"
        )
    return manuals


def build_cases(sources: list[SourceCommit], manuals: list[ManualCommit]) -> list[EvalCase]:
    source_by_commit = {source.source_commit: source for source in sources}
    source_by_title: dict[str, list[SourceCommit]] = {}
    for source in sources:
        source_by_title.setdefault(normalize_commit_title(source.source_title), []).append(source)

    manual_by_source: dict[str, ManualCommit] = {}

    def map_manual(source_commit: str, manual: ManualCommit) -> None:
        existing = manual_by_source.get(source_commit)
        if existing and existing.commit != manual.commit:
            source = source_by_commit[source_commit]
            if source.target_commits:
                preferred = source.target_commits[0]
                if manual.commit == preferred:
                    LOGGER.warning(
                        "source commit %s maps to multiple manual commits; "
                        "using Excel target commit %s instead of %s",
                        source_commit,
                        manual.commit,
                        existing.commit,
                    )
                    manual_by_source[source_commit] = manual
                else:
                    LOGGER.warning(
                        "source commit %s maps to multiple manual commits; "
                        "keeping %s and ignoring %s",
                        source_commit,
                        existing.commit,
                        manual.commit,
                    )
                return
            raise EvalError(
                f"source commit {source_commit} maps to two manual commits: "
                f"{existing.commit} and {manual.commit}"
            )
        manual_by_source[source_commit] = manual

    manual_by_commit = {manual.commit: manual for manual in manuals}
    for source in sources:
        for target_commit in source.target_commits:
            manual = manual_by_commit.get(target_commit)
            if manual:
                map_manual(source.source_commit, manual)

    for manual in manuals:
        candidates = [ref for ref in manual.source_refs if ref in source_by_commit]
        if not candidates:
            title_matches = source_by_title.get(normalize_commit_title(manual.title), [])
            if len(title_matches) == 1:
                candidates = [title_matches[0].source_commit]
        for source_commit in candidates:
            source = source_by_commit[source_commit]
            if source.target_commits and manual.commit not in source.target_commits:
                continue
            map_manual(source_commit, manual)

    next_manual: ManualCommit | None = None
    cases_reversed: list[EvalCase] = []
    for source in reversed(sources):
        manual = manual_by_source.get(source.source_commit)
        if manual:
            next_manual = manual
            cases_reversed.append(
                EvalCase(
                    source=source,
                    manual_pr_index=manual.index,
                    manual_commit=manual.commit,
                    case_baseline=manual.parent,
                    expected_behavior="manual_backport",
                )
            )
        elif next_manual:
            cases_reversed.append(
                EvalCase(
                    source=source,
                    manual_pr_index=None,
                    manual_commit="",
                    case_baseline=next_manual.parent,
                    expected_behavior="equivalent_before_pr",
                )
            )
        else:
            cases_reversed.append(
                EvalCase(
                    source=source,
                    manual_pr_index=None,
                    manual_commit="",
                    case_baseline=manuals[-1].commit,
                    expected_behavior="equivalent_before_pr",
                )
            )
    return list(reversed(cases_reversed))


def validate_discovery(
    config: Config,
    sources: list[SourceCommit],
    manuals: list[ManualCommit],
    cases: list[EvalCase],
    input_count: int,
) -> None:
    mapped = [case for case in cases if case.manual_commit]
    for case in mapped:
        expected_parent = git(config.target_repo, "rev-parse", f"{case.manual_commit}^")
        if case.case_baseline != expected_parent:
            raise EvalError(f"incorrect mapped baseline for {case.source.source_commit}")
    next_mapped_case: EvalCase | None = None
    for case in reversed(cases):
        if case.manual_commit:
            next_mapped_case = case
            continue
        expected_baseline = (
            next_mapped_case.case_baseline if next_mapped_case else config.last_pr_commit
        )
        if case.case_baseline != expected_baseline:
            raise EvalError(
                f"incorrect unmapped baseline for {case.source.source_commit}: "
                f"{case.case_baseline}, expected {expected_baseline}"
            )


def discovery_payload(
    config: Config,
    sources: list[SourceCommit],
    manuals: list[ManualCommit],
    cases: list[EvalCase],
    input_count: int,
    source_input_mode: str,
) -> dict[str, Any]:
    return {
        "source_repo": str(config.source_repo),
        "source_excel": str(config.source_excel) if config.source_excel else "",
        "preprocessed_source_excel": (
            str(config.preprocessed_source_excel) if config.preprocessed_source_excel else ""
        ),
        "source_input_mode": source_input_mode,
        "target_repo": str(config.target_repo),
        "manual_patch_dir": str(config.manual_patch_dir) if config.manual_patch_dir else "",
        "manual_input_mode": config.manual_input_mode,
        "pr_url": config.pr_url,
        "first_pr_commit": config.first_pr_commit,
        "last_pr_commit": config.last_pr_commit,
        "pr_baseline": config.pr_baseline,
        "input_count": input_count,
        "unique_source_count": len(sources),
        "manual_commit_count": len(manuals),
        "mapped_count": sum(bool(case.manual_commit) for case in cases),
        "cases": [
            {
                **asdict(case.source),
                "manual_pr_index": case.manual_pr_index,
                "manual_commit": case.manual_commit,
                "case_baseline": case.case_baseline,
                "expected_behavior": case.expected_behavior,
            }
            for case in cases
        ],
    }


def ensure_clean_target(config: Config) -> None:
    status = git(config.target_repo, "status", "--porcelain", "--untracked-files=normal")
    if status:
        raise EvalError(
            "target repository must be clean before evaluation:\n"
            f"{status}"
        )


def is_eval_temp_branch(branch: str) -> bool:
    return branch.startswith("eval/") and "backport-eval" in branch


def current_repo_state(config: Config) -> dict[str, str]:
    branch = git(config.target_repo, "symbolic-ref", "--quiet", "--short", "HEAD", check=False)
    if branch and is_eval_temp_branch(branch):
        raise EvalError(
            "target repository is currently on an eval temp branch from an unfinished "
            f"or uncleaned run: {branch}. Resume or recover that run before starting "
            f"{config.eval_name}."
        )
    return {
        "eval_name": config.eval_name,
        "branch": branch,
        "head": git(config.target_repo, "rev-parse", "HEAD"),
        "temp_branch": config.temp_branch,
        "target_repo": str(config.target_repo),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }


def write_recovery(config: Config, state: dict[str, str]) -> None:
    config.recovery_file.write_text(json.dumps(state, indent=2) + "\n", encoding="utf-8")


def write_json_atomic(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_name(f".{path.name}.tmp-{os.getpid()}")
    temp_path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    temp_path.replace(path)


def read_json(path: Path) -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise EvalError(f"failed to read JSON checkpoint {path}: {exc}") from exc


def discovery_signature(payload: dict[str, Any]) -> dict[str, Any]:
    signature = {
        "source_input_mode": payload.get("source_input_mode", ""),
        "source_excel": payload.get("source_excel", ""),
        "manual_patch_dir": payload.get("manual_patch_dir", ""),
        "manual_input_mode": payload.get("manual_input_mode", ""),
        "last_pr_commit": payload.get("last_pr_commit") or payload.get("pr_head", ""),
        "pr_baseline": payload["pr_baseline"],
        "input_count": payload["input_count"],
        "unique_source_count": payload["unique_source_count"],
        "manual_commit_count": payload["manual_commit_count"],
        "mapped_count": payload["mapped_count"],
        "cases": [
            {
                "describe_index": case["describe_index"],
                "source_commit": case["source_commit"],
                "manual_commit": case["manual_commit"],
                "case_baseline": case["case_baseline"],
                "expected_behavior": case["expected_behavior"],
            }
            for case in payload["cases"]
        ],
    }
    first_pr_commit = payload.get("first_pr_commit")
    if first_pr_commit:
        signature["first_pr_commit"] = first_pr_commit
    return signature


def validate_resume_discovery(run_dir: Path, payload: dict[str, Any]) -> None:
    discovery_path = run_dir / "discovery.json"
    previous = read_json(discovery_path)
    if discovery_signature(previous) != discovery_signature(payload):
        raise EvalError(
            f"resume input no longer matches the original run discovery: {discovery_path}"
        )


def case_checkpoint_paths(case_dir: Path) -> tuple[Path, Path]:
    return case_dir / "result.json", case_dir / "conflict_detail.json"


def load_completed_case(case_dir: Path, case: EvalCase) -> tuple[dict[str, Any], dict[str, Any] | None] | None:
    result_path, detail_path = case_checkpoint_paths(case_dir)
    if not result_path.exists():
        return None
    row = read_json(result_path)
    if row.get("source_commit") != case.source.source_commit:
        raise EvalError(f"case checkpoint source mismatch: {result_path}")
    if row.get("case_baseline") != case.case_baseline:
        raise EvalError(f"case checkpoint baseline mismatch: {result_path}")
    if row.get("decision") == "failed":
        return None
    detail = read_json(detail_path) if detail_path.exists() else None
    return row, detail


def next_attempt_dir(case_dir: Path) -> Path:
    attempt = 1
    while True:
        path = case_dir / f"attempt_{attempt:03d}"
        if not path.exists():
            path.mkdir(parents=True)
            return path
        attempt += 1


def write_case_checkpoint(
    case_dir: Path,
    row: dict[str, Any],
    detail: dict[str, Any] | None,
) -> None:
    result_path, detail_path = case_checkpoint_paths(case_dir)
    write_json_atomic(result_path, row)
    if detail:
        write_json_atomic(detail_path, detail)
    else:
        detail_path.unlink(missing_ok=True)


def write_run_checkpoint(
    run_dir: Path,
    rows: list[dict[str, Any]],
    total_cases: int,
    output_path: Path,
) -> None:
    completed = [row for row in rows if row.get("decision") != "failed"]
    failed = [row for row in rows if row.get("decision") == "failed"]
    write_json_atomic(
        run_dir / "checkpoint.json",
        {
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "total_cases": total_cases,
            "checkpointed_cases": len(rows),
            "completed_cases": len(completed),
            "failed_cases_to_retry": len(failed),
            "completed_source_commits": [row["source_commit"] for row in completed],
            "failed_source_commits": [row["source_commit"] for row in failed],
            "output_path": str(output_path),
        },
    )


def restore_target(config: Config, state: dict[str, str]) -> None:
    LOGGER.info(
        "RESTORE: target repository branch=%s head=%s",
        state.get("branch") or "(detached)",
        state["head"],
    )
    git(config.target_repo, "reset", "--hard")
    git(config.target_repo, "clean", "-fd")
    if state.get("branch"):
        if is_eval_temp_branch(state["branch"]):
            LOGGER.warning(
                "RESTORE: recorded branch is an eval temp branch; detaching at saved head branch=%s head=%s",
                state["branch"],
                state["head"],
            )
            git(config.target_repo, "switch", "--detach", state["head"])
        else:
            branch_exists = git(
                config.target_repo,
                "rev-parse",
                "--verify",
                f"{state['branch']}^{{commit}}",
                check=False,
            )
            if branch_exists:
                git(config.target_repo, "switch", state["branch"])
                git(config.target_repo, "reset", "--hard", state["head"])
            else:
                LOGGER.warning(
                    "RESTORE: recorded branch no longer exists; detaching at saved head branch=%s head=%s",
                    state["branch"],
                    state["head"],
                )
                git(config.target_repo, "switch", "--detach", state["head"])
    else:
        git(config.target_repo, "switch", "--detach", state["head"])
    recovery_temp_branch = state.get("temp_branch") or config.temp_branch
    git(config.target_repo, "branch", "-D", recovery_temp_branch, check=False)
    config.recovery_file.unlink(missing_ok=True)


def recover_if_needed(config: Config) -> None:
    if not config.recovery_file.exists():
        return
    state = json.loads(config.recovery_file.read_text(encoding="utf-8"))
    if state.get("target_repo") != str(config.target_repo):
        raise EvalError(f"unexpected recovery record: {config.recovery_file}")
    if state.get("eval_name") and state["eval_name"] != config.eval_name:
        raise EvalError(
            f"recovery record belongs to {state['eval_name']}, not {config.eval_name}: "
            f"{config.recovery_file}"
        )
    LOGGER.warning("recovering target repository from an interrupted previous run")
    restore_target(config, state)


def ensure_no_foreign_recovery(config: Config) -> None:
    for path in SCRIPT_DIR.glob("*_recovery.json"):
        if path.resolve() == config.recovery_file.resolve():
            continue
        try:
            state = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        if state.get("target_repo") == str(config.target_repo):
            raise EvalError(
                "another evaluation has an unfinished recovery record for the same "
                f"target repository: {path}. Resume or recover that run before "
                f"starting {config.eval_name}."
            )


def reset_case(config: Config, baseline: str) -> None:
    LOGGER.info(
        "CASE RESET: recreate/switch branch=%s at baseline=%s",
        config.temp_branch,
        baseline,
    )
    # Recreate the run-owned branch for every independent case. This both
    # guarantees the correct baseline and self-heals if an external command
    # deleted the temporary branch between cases.
    git(config.target_repo, "switch", "-C", config.temp_branch, baseline)
    git(config.target_repo, "reset", "--hard", baseline)
    git(config.target_repo, "clean", "-fd")
    if git(config.target_repo, "rev-parse", "HEAD") != baseline:
        raise EvalError(f"failed to reset target repository to case baseline {baseline}")


def cvekit_command(config: Config, report_or_raw: Path, action: str, apply: str = "") -> list[str]:
    cmd = [
        str(config.cvekit),
        "--action",
        "backport-batch",
        "--backport-config",
        str(report_or_raw),
        "--commit-sort",
        SINGLE_CASE_COMMIT_SORT,
        "--backport-engine",
        "mystique",
        "--format-mode",
        "changed",
        "--debug",
        "--json",
    ]
    if config.llm_provider:
        cmd.extend(["--llm-provider", config.llm_provider])
    if config.llm_base_url:
        cmd.extend(["--llm-base-url", config.llm_base_url])
    if config.llm_model_name:
        cmd.extend(["--llm-model-name", config.llm_model_name])
    if config.api_key:
        cmd.extend(["--api-key", config.api_key])
    if action == "execute":
        cmd.append("--execute")
    elif action == "apply":
        cmd.extend(
            [
                "--apply",
                apply,
                "--signer-name",
                config.signer_name,
                "--signer-email",
                config.signer_email,
            ]
        )
    elif action != "raw":
        raise EvalError(f"unsupported cvekit action: {action}")
    return cmd


def write_raw_config(config: Config, case: EvalCase, case_dir: Path) -> Path:
    raw_path = case_dir / "backport-batch.yml"
    data = {
        "project": "linux",
        "project_dir": str(config.source_repo),
        "source_branch": config.source_branch,
        "target_path": str(config.target_repo),
        "target_release": config.temp_branch,
        "target_branch": config.temp_branch,
        "patch_dataset_dir": str(case_dir / "patch-dataset"),
        "backport_engine": "mystique",
        "format_mode": "changed",
        "commit_sort": SINGLE_CASE_COMMIT_SORT,
        "signer_name": config.signer_name,
        "signer_email": config.signer_email,
        "commits": [
            {
                "commit": case.source.source_commit,
                "tag": f"{case.source.describe_index:03d}_{case.source.source_commit[:12]}",
                "target_branch": config.temp_branch,
            }
        ],
    }
    commits = data["commits"]
    if len(commits) != 1 or "commit_title" in commits[0]:
        raise EvalError(
            "single-case optimization requires exactly one exact commit SHA "
            "without commit_title"
        )
    if config.llm_provider:
        data["llm_provider"] = config.llm_provider
    if config.llm_base_url:
        data["llm_base_url"] = config.llm_base_url
    if config.llm_model_name:
        data["llm_model_name"] = config.llm_model_name
    raw_path.write_text(yaml.safe_dump(data, allow_unicode=True, sort_keys=False), encoding="utf-8")
    return raw_path


def load_report_item(report_path: Path) -> dict[str, Any]:
    data = yaml.safe_load(report_path.read_text(encoding="utf-8"))
    commits = data.get("commits", []) if isinstance(data, dict) else []
    if len(commits) != 1:
        raise EvalError(f"expected one report item in {report_path}, got {len(commits)}")
    return commits[0]


def patch_id(patch: str) -> str:
    if not patch.strip():
        return ""
    completed = run(["git", "patch-id", "--stable"], input_text=patch, check=False)
    return completed.stdout.split()[0] if completed.returncode == 0 and completed.stdout else ""


def diff(repo: Path, old: str, new: str) -> str:
    return git(repo, "diff", "--binary", "--full-index", old, new)


def copy_text(path_value: Any) -> str:
    path = Path(str(path_value or "")).expanduser()
    if not path_value or not path.is_file():
        return ""
    return path.read_text(encoding="utf-8", errors="replace")


def run_case(config: Config, case: EvalCase, case_dir: Path) -> tuple[dict[str, Any], dict[str, Any] | None]:
    reset_case(config, case.case_baseline)
    start_head = git(config.target_repo, "rev-parse", "HEAD")
    raw_path = write_raw_config(config, case, case_dir)
    report_path = Path(f"{raw_path}.report.yml")

    LOGGER.info(
        "CASE RAW: checking merged/conflict status source=%s report=%s",
        case.source.source_commit[:12],
        report_path,
    )
    run(
        cvekit_command(config, raw_path, "raw"),
        cwd=config.cvekit_workdir,
        log_path=case_dir / "raw.log",
    )
    if not report_path.exists():
        raise EvalError(f"cvekit did not create report: {report_path}")
    initial_item = load_report_item(report_path)
    initial_merged = bool(initial_item.get("merged_in_target"))
    initial_conflict = initial_item.get("has_conflict")
    LOGGER.info(
        "CASE RAW RESULT: source=%s merged_in_target=%s has_conflict=%s method=%s",
        case.source.source_commit[:12],
        initial_merged,
        initial_conflict,
        initial_item.get("conflict_check_method") or "",
    )

    decision = "already_merged" if initial_merged else "failed"
    execution_status = str(initial_item.get("status") or "")
    apply_success = False
    final_item = initial_item
    warning = ""

    if not initial_merged:
        if initial_conflict:
            LOGGER.info(
                "CASE EXECUTE: conflict detected; invoking Mystique source=%s",
                case.source.source_commit[:12],
            )
        else:
            LOGGER.info(
                "CASE EXECUTE: no conflict; applying source commit directly source=%s",
                case.source.source_commit[:12],
            )
        execute = run(
            cvekit_command(config, report_path, "execute"),
            cwd=config.cvekit_workdir,
            check=False,
            log_path=case_dir / "execute.log",
        )
        final_item = load_report_item(report_path)
        execution_status = str(final_item.get("status") or "")
        backported_patch = str(final_item.get("backported_patch_path") or "")
        apply_method = str(final_item.get("apply_method") or "")
        fallback_error = str(final_item.get("fallback_error") or "")
        warning = str(final_item.get("warning") or "")
        equivalent = bool(final_item.get("empty_patch") or final_item.get("equivalent_exists"))
        if equivalent:
            decision = "need_not_ported"
            apply_success = True
            LOGGER.info(
                "CASE DECISION: Mystique/equivalence check found no port needed source=%s",
                case.source.source_commit[:12],
            )
        elif initial_conflict and backported_patch:
            decision = "mystique_patch"
            LOGGER.info(
                "CASE APPLY: Mystique generated patch; applying patch=%s",
                backported_patch,
            )
            applied = run(
                cvekit_command(config, report_path, "apply", backported_patch),
                cwd=config.cvekit_workdir,
                check=False,
                log_path=case_dir / "apply.log",
            )
            apply_success = applied.returncode == 0 and git(config.target_repo, "rev-parse", "HEAD") != start_head
            LOGGER.info("CASE APPLY RESULT: source=%s success=%s", case.source.source_commit[:12], apply_success)
        elif not initial_conflict:
            decision = "direct_apply"
            apply_success = execution_status == "success" and git(config.target_repo, "rev-parse", "HEAD") != start_head
            LOGGER.info(
                "CASE DIRECT APPLY RESULT: source=%s success=%s method=%s",
                case.source.source_commit[:12],
                apply_success,
                apply_method or "unknown",
            )
        else:
            LOGGER.warning(
                "CASE EXECUTE RESULT: conflict path produced no Mystique patch source=%s status=%s error=%s",
                case.source.source_commit[:12],
                execution_status,
                final_item.get("error") or "",
            )
    else:
        LOGGER.info(
            "CASE DECISION: source already merged in case baseline source=%s",
            case.source.source_commit[:12],
        )

    final_head = git(config.target_repo, "rev-parse", "HEAD")
    changed = final_head != start_head or bool(git(config.target_repo, "status", "--porcelain"))
    no_change_skip = (
        not changed
        and str(final_item.get("status") or "").lower() == "skipped"
        and any(
            marker in str(final_item.get("error") or "").lower()
            for marker in ("equivalent", "已包含", "已生效", "未产生变更")
        )
    )
    if no_change_skip:
        decision = "need_not_ported"
        apply_success = True
    equivalent_detected = decision in {"already_merged", "need_not_ported"}
    missed_equivalence = case.expected_behavior == "equivalent_before_pr" and changed
    manual_patch = ""
    ai_patch = diff(config.target_repo, case.case_baseline, final_head) if changed else ""
    tree_match_manual: bool | None = None
    patch_id_match_manual: bool | None = None
    if case.manual_commit:
        manual_patch = diff(config.target_repo, case.case_baseline, case.manual_commit)
        tree_match_manual = (
            git(config.target_repo, "rev-parse", f"{final_head}^{{tree}}")
            == git(config.target_repo, "rev-parse", f"{case.manual_commit}^{{tree}}")
        )
        patch_id_match_manual = patch_id(ai_patch) == patch_id(manual_patch)

    row = {
        "describe_index": case.source.describe_index,
        "source_excel_rows": ",".join(map(str, case.source.excel_rows)),
        "source_commit": case.source.source_commit,
        "source_title": case.source.source_title,
        "git_describe": case.source.git_describe,
        "manual_pr_index": case.manual_pr_index,
        "manual_commit": case.manual_commit,
        "case_baseline": case.case_baseline,
        "expected_behavior": case.expected_behavior,
        "initial_merged_in_target": initial_merged,
        "initial_has_conflict": initial_conflict,
        "decision": decision,
        "execution_status": execution_status,
        "apply_success": apply_success,
        "apply_method": str(final_item.get("apply_method") or ""),
        "fallback_error": str(final_item.get("fallback_error") or ""),
        "warning": warning,
        "mystique_status": final_item.get("status") if initial_conflict else "",
        "equivalent_detected": equivalent_detected,
        "missed_equivalence": missed_equivalence,
        "tree_match_manual": tree_match_manual,
        "patch_id_match_manual": patch_id_match_manual,
        "error_summary": final_item.get("error") or "",
    }
    conflict_detail = None
    if initial_conflict:
        conflict_detail = {
            "source_commit": case.source.source_commit,
            "manual_commit": case.manual_commit,
            "manual_patch": manual_patch,
            "mystique_backported_patch": copy_text(final_item.get("backported_patch_path")),
            "mystique_backported_log_path": str(final_item.get("logfile") or ""),
        }
    LOGGER.info(
        "CASE COMPLETE: source=%s decision=%s changed=%s apply_success=%s tree_match_manual=%s patch_id_match_manual=%s",
        case.source.source_commit[:12],
        decision,
        changed,
        apply_success,
        tree_match_manual,
        patch_id_match_manual,
    )
    return row, conflict_detail


RESULT_COLUMNS = [
    "describe_index",
    "source_excel_rows",
    "source_commit",
    "source_title",
    "git_describe",
    "manual_pr_index",
    "manual_commit",
    "case_baseline",
    "expected_behavior",
    "initial_merged_in_target",
    "initial_has_conflict",
    "decision",
    "execution_status",
    "apply_success",
    "apply_method",
    "fallback_error",
    "warning",
    "mystique_status",
    "equivalent_detected",
    "missed_equivalence",
    "tree_match_manual",
    "patch_id_match_manual",
    "error_summary",
]


def append_chunked_row(sheet: Any, row: dict[str, Any], columns: list[str]) -> None:
    chunks_by_column: list[list[Any]] = []
    max_chunks = 1
    for column in columns:
        value = row.get(column, "")
        if isinstance(value, str) and len(value) > EXCEL_CELL_LIMIT:
            chunks = [
                value[index : index + EXCEL_CELL_LIMIT]
                for index in range(0, len(value), EXCEL_CELL_LIMIT)
            ]
        else:
            chunks = [value]
        chunks_by_column.append(chunks)
        max_chunks = max(max_chunks, len(chunks))
    for chunk_index in range(max_chunks):
        sheet.append(
            [
                chunks[chunk_index] if chunk_index < len(chunks) else ""
                for chunks in chunks_by_column
            ]
        )


def ratio(rows: list[dict[str, Any]], key: str, predicate: str | None = None) -> str:
    selected = rows if predicate is None else [row for row in rows if row.get("expected_behavior") == predicate]
    if not selected:
        return "0/0"
    passed = sum(row.get(key) is True for row in selected)
    return f"{passed}/{len(selected)} ({passed / len(selected):.2%})"


def write_workbook(
    config: Config,
    output_path: Path,
    rows: list[dict[str, Any]],
    conflict_details: list[dict[str, Any]],
    payload: dict[str, Any],
    started_at: str,
    ended_at: str,
) -> None:
    workbook = openpyxl.Workbook()
    results = workbook.active
    results.title = "Results"
    results.append(RESULT_COLUMNS)
    for row in rows:
        append_chunked_row(results, row, RESULT_COLUMNS)

    conflicts = workbook.create_sheet("ConflictDetails")
    conflict_columns = [
        "source_commit",
        "manual_commit",
        "manual_patch",
        "mystique_backported_patch",
        "mystique_backported_log_path",
    ]
    conflicts.append(conflict_columns)
    for detail in conflict_details:
        append_chunked_row(conflicts, detail, conflict_columns)

    run_info = workbook.create_sheet("RunInfo")
    run_info.append(["key", "value"])
    manual_rows = [row for row in rows if row["manual_commit"]]
    manual_conflict_rows = [row for row in manual_rows if row.get("initial_has_conflict") is True]
    manual_mystique_rows = [row for row in manual_rows if row.get("decision") == "mystique_patch"]
    info = {
        "source_repo": str(config.source_repo),
        "source_excel": str(config.source_excel) if config.source_excel else "",
        "preprocessed_source_excel": (
            str(config.preprocessed_source_excel) if config.preprocessed_source_excel else ""
        ),
        "source_input_mode": payload.get("source_input_mode", ""),
        "target_repo": str(config.target_repo),
        "manual_patch_dir": str(config.manual_patch_dir) if config.manual_patch_dir else "",
        "manual_input_mode": config.manual_input_mode,
        "PR URL": config.pr_url,
        "PR first commit": config.first_pr_commit,
        "PR last commit": config.last_pr_commit,
        "derived PR baseline": config.pr_baseline,
        "source branch head": git(config.source_repo, "rev-parse", config.source_branch),
        "input count": payload["input_count"],
        "unique source commit count": payload["unique_source_count"],
        "PR commit count": payload["manual_commit_count"],
        "mapping count": payload["mapped_count"],
        "llm_provider": config.llm_provider,
        "llm_base_url": config.llm_base_url,
        "llm_model_name": config.llm_model_name,
        "Mystique parameters": (
            f"outer order: {DISCOVERY_COMMIT_SORT}; single-case cvekit: "
            f"--commit-sort {SINGLE_CASE_COMMIT_SORT} --backport-engine mystique "
            "--format-mode changed"
        ),
        "started_at": started_at,
        "ended_at": ended_at,
        "direct apply success rate": ratio(
            [row for row in rows if row["decision"] == "direct_apply"], "apply_success"
        ),
        "Mystique patch success rate": ratio(
            [row for row in rows if row["decision"] == "mystique_patch"], "apply_success"
        ),
        "equivalence detection success rate": ratio(rows, "equivalent_detected", "equivalent_before_pr"),
        "missed_equivalence count": sum(bool(row.get("missed_equivalence")) for row in rows),
        "mapped manual case count": len(manual_rows),
        "mapped manual conflict count": len(manual_conflict_rows),
        "mapped manual Mystique patch count": len(manual_mystique_rows),
        "manual tree match rate": ratio(manual_rows, "tree_match_manual"),
        "manual patch-id match rate": ratio(manual_rows, "patch_id_match_manual"),
    }
    for key, value in info.items():
        run_info.append([key, value])
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    workbook.save(output_path)


def main() -> int:
    setup_logging()
    args = parse_args()
    config = build_config(args)
    _run_lock = acquire_run_lock(config)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    config.temp_branch = f"{config.temp_branch}-{stamp.lower()}"
    resume_mode = bool(args.resume)
    run_dir = (
        Path(args.resume).expanduser().resolve()
        if resume_mode
        else config.log_root / f"{config.eval_name}_{stamp}"
    )
    cases_dir = run_dir / "cases"
    if resume_mode:
        required = [
            run_dir / "discovery.json",
            run_dir / "run_state.json",
            run_dir / "checkpoint.json",
            cases_dir,
        ]
        missing = [str(path) for path in required if not path.exists()]
        if missing:
            raise EvalError(
                "resume requires a run created by the checkpoint-enabled evaluator; "
                f"missing: {missing}"
            )
    else:
        cases_dir.mkdir(parents=True, exist_ok=False)
    attach_run_log(run_dir / "run.log")
    LOGGER.info("run directory: %s", run_dir)
    LOGGER.info(
        "RUN LOCK: acquired %s; temporary branch=%s; resume=%s",
        config.lock_file,
        config.temp_branch,
        resume_mode,
    )
    recover_if_needed(config)
    ensure_no_foreign_recovery(config)
    required_paths = [config.source_repo, config.target_repo, config.cvekit]
    if config.source_excel is not None:
        required_paths.append(config.source_excel)
    if config.manual_patch_dir is not None:
        required_paths.append(config.manual_patch_dir)
    for path in required_paths:
        if not path.exists():
            raise EvalError(f"required path does not exist: {path}")
    git_verify(config.source_repo, config.source_branch)
    git_verify(config.target_repo, config.first_pr_commit)
    git_verify(config.target_repo, config.last_pr_commit)
    git_verify(config.target_repo, config.pr_baseline)

    manuals = read_manual_commits(config)
    if config.source_excel is not None:
        LOGGER.info("reading and sorting Excel source commits")
        preprocessed_excel_path = run_dir / f"{config.source_excel.stem}.with_commit_hash.xlsx"
        sources, input_count, source_input_mode = read_source_commits_from_excel(
            config,
            manuals,
            preprocessed_excel_path,
        )
    else:
        LOGGER.info("extracting and sorting source commits from PR commit messages")
        sources, input_count, source_input_mode = read_source_commits_from_pr_messages(config, manuals)
    cases = build_cases(sources, manuals)
    validate_discovery(config, sources, manuals, cases, input_count)
    payload = discovery_payload(config, sources, manuals, cases, input_count, source_input_mode)
    if resume_mode:
        validate_resume_discovery(run_dir, payload)
    else:
        write_json_atomic(run_dir / "discovery.json", payload)
    LOGGER.info(
        "discovery validated: source_mode=%s manual_mode=%s inputs=%d unique=%d manual=%d mapped=%d",
        source_input_mode,
        config.manual_input_mode,
        input_count,
        len(sources),
        len(manuals),
        payload["mapped_count"],
    )
    LOGGER.info(
        "SORT POLICY: outer evaluation order=%s; single-case cvekit sort=%s "
        "(one exact SHA, no commit_title; avoids full source-branch title scan)",
        DISCOVERY_COMMIT_SORT,
        SINGLE_CASE_COMMIT_SORT,
    )
    if args.discover_only:
        LOGGER.info("discover-only result written: %s", run_dir / "discovery.json")
        print(json.dumps(payload, ensure_ascii=False, indent=2))
        return 0

    ensure_clean_target(config)
    state = current_repo_state(config)
    write_recovery(config, state)
    LOGGER.info(
        "RECOVERY RECORD: saved original branch=%s head=%s path=%s",
        state.get("branch") or "(detached)",
        state["head"],
        config.recovery_file,
    )
    run_state_path = run_dir / "run_state.json"
    if resume_mode:
        run_state = read_json(run_state_path)
        started_at = str(run_state["started_at"])
        output_path = Path(str(run_state["output_path"])).expanduser().resolve()
    else:
        started_at = datetime.now(timezone.utc).isoformat()
        output_path = config.output_dir / f"{config.eval_name}_{stamp}.xlsx"
        write_json_atomic(
            run_state_path,
            {
                "started_at": started_at,
                "output_path": str(output_path),
                "run_dir": str(run_dir),
            },
        )
    rows: list[dict[str, Any]] = []
    conflicts: list[dict[str, Any]] = []
    selected_cases = cases[: config.case_limit] if config.case_limit else cases
    try:
        LOGGER.info("WORKSPACE: creating/resetting temporary branch=%s", config.temp_branch)
        git(config.target_repo, "switch", "-C", config.temp_branch, config.pr_baseline)
        for index, case in enumerate(selected_cases, start=1):
            case_dir = cases_dir / f"{case.source.describe_index:03d}_{safe_name(case.source.source_title)}"
            case_dir.mkdir(parents=True, exist_ok=True)
            completed = load_completed_case(case_dir, case) if resume_mode else None
            if completed:
                row, detail = completed
                rows.append(row)
                if detail:
                    conflicts.append(detail)
                LOGGER.info(
                    "RESUME SKIP: case %d/%d source=%s decision=%s",
                    index,
                    len(selected_cases),
                    case.source.source_commit[:12],
                    row.get("decision"),
                )
                continue
            LOGGER.info(
                "case %d/%d source=%s baseline=%s expected=%s",
                index,
                len(selected_cases),
                case.source.source_commit[:12],
                case.case_baseline[:12],
                case.expected_behavior,
            )
            attempt_dir = next_attempt_dir(case_dir)
            LOGGER.info("CASE ATTEMPT: source=%s path=%s", case.source.source_commit[:12], attempt_dir)
            try:
                row, detail = run_case(config, case, attempt_dir)
            except Exception as exc:
                LOGGER.exception("case failed: %s", case.source.source_commit)
                row = {
                    **{
                        "describe_index": case.source.describe_index,
                        "source_excel_rows": ",".join(map(str, case.source.excel_rows)),
                        "source_commit": case.source.source_commit,
                        "source_title": case.source.source_title,
                        "git_describe": case.source.git_describe,
                        "manual_pr_index": case.manual_pr_index,
                        "manual_commit": case.manual_commit,
                        "case_baseline": case.case_baseline,
                        "expected_behavior": case.expected_behavior,
                    },
                    "decision": "failed",
                    "execution_status": "failed",
                    "error_summary": str(exc),
                }
                detail = None
            write_case_checkpoint(case_dir, row, detail)
            rows.append(row)
            if detail:
                conflicts.append(detail)
            ended_at = datetime.now(timezone.utc).isoformat()
            write_run_checkpoint(run_dir, rows, len(selected_cases), output_path)
            write_workbook(config, output_path, rows, conflicts, payload, started_at, ended_at)
            LOGGER.info(
                "CHECKPOINT: case=%d/%d result=%s workbook=%s",
                index,
                len(selected_cases),
                case_dir / "result.json",
                output_path,
            )
        ended_at = datetime.now(timezone.utc).isoformat()
        write_run_checkpoint(run_dir, rows, len(selected_cases), output_path)
        write_workbook(config, output_path, rows, conflicts, payload, started_at, ended_at)
        LOGGER.info("evaluation workbook written: %s", output_path)
    finally:
        restore_target(config, state)
        LOGGER.info("persistent case logs kept at %s", cases_dir)
    print(output_path)
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except EvalError as exc:
        print(f"error: {exc}", file=sys.stderr)
        raise SystemExit(1)
