#!/usr/bin/env python3
"""Layered similarity evaluation for ai_backport_eval workbooks.

Workbook contract:
- Read `decision`, `expected_behavior`, and `case_baseline` from the
  `Results` sheet.
- Read `manual_patch` and `mystique_backported_patch` from the
  `ConflictDetails` sheet.
- Read `source_repo` and `target_repo` from the `RunInfo` sheet when present.
- Join sheets by `source_commit`.
- Reassemble continuation rows created by Excel's cell-length limit.
- Evaluate `expected_behavior=manual_backport` and
  `decision=mystique_patch` by default.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import subprocess
from collections import Counter
from dataclasses import asdict, dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable

import openpyxl


WORKBOOK_CONTRACT = """\
ai_backport_eval workbook contract:
  - Read decision, expected_behavior, and case_baseline from Results.
  - Read manual_patch and mystique_backported_patch from ConflictDetails.
  - Read source_repo and target_repo from RunInfo when present.
  - Join sheets by source_commit.
  - Reassemble continuation rows created by Excel's cell-length limit.
  - Evaluate expected_behavior=manual_backport and decision=mystique_patch by default.
"""


DIFF_START_RE = re.compile(r"^diff --git ", re.MULTILINE)
HUNK_RE = re.compile(r"^@@[^@]*@@\s*(.*)$")
TOKEN_RE = re.compile(r"[A-Za-z_][A-Za-z0-9_]*|\b\d+\b")


@dataclass
class PatchFeatures:
    valid: bool
    files: list[str]
    additions: int
    deletions: int
    changed_lines: list[str]
    tokens: list[str]
    hunk_labels: list[str]
    normalized_patch: str
    patch_id: str


def truthy(value: Any) -> bool | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"true", "1", "yes"}:
        return True
    if text in {"false", "0", "no"}:
        return False
    return None


def sheet_rows(sheet: Any) -> list[dict[str, Any]]:
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        return []
    headers = [str(value or "").strip() for value in values[0]]
    return [dict(zip(headers, row)) for row in values[1:]]


def reassemble_chunked_rows(rows: list[dict[str, Any]], key: str) -> list[dict[str, Any]]:
    """Join continuation rows emitted by append_chunked_row."""
    assembled: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    for row in rows:
        if row.get(key):
            current = dict(row)
            assembled.append(current)
            continue
        if current is None:
            continue
        for column, value in row.items():
            if isinstance(value, str) and value:
                current[column] = f"{current.get(column) or ''}{value}"
    return assembled


def run_info_values(workbook: Any) -> dict[str, str]:
    """读取 RunInfo 中的仓库上下文；旧表格没有这些字段时保持为空。"""
    if "RunInfo" not in workbook.sheetnames:
        return {"source_repo": "", "target_repo": ""}
    values: dict[str, str] = {}
    for row in workbook["RunInfo"].iter_rows(min_row=2, values_only=True):
        key = row[0] if row else None
        value = row[1] if len(row) > 1 else ""
        if key:
            values[str(key)] = "" if value is None else str(value)
    return {
        "source_repo": values.get("source_repo", ""),
        "target_repo": values.get("target_repo", ""),
    }


def stable_patch_id(patch: str) -> str:
    if not patch.strip():
        return ""
    # Git stable patch-id 用来判断“diff 语义指纹”是否一致；它会弱化行号、
    # 空白和文件顺序等噪音，比直接比较 patch 文本更适合作为强相似指标。
    completed = subprocess.run(
        ["git", "patch-id", "--stable"],
        input=patch,
        text=True,
        capture_output=True,
        check=False,
    )
    return completed.stdout.split()[0] if completed.returncode == 0 and completed.stdout else ""


def normalize_code_line(line: str) -> str:
    # coverage/precision 只关心变更行内容本身，先去掉空白差异；
    # 行首的 +/- 会在调用方保留，用来区分“新增”和“删除”。
    return re.sub(r"\s+", "", line).strip()


def normalize_patch(patch: str) -> str:
    match = DIFF_START_RE.search(patch)
    body = patch[match.start() :] if match else patch
    normalized: list[str] = []
    for line in body.splitlines():
        # index 行和 hunk 行号通常只反映生成环境，不代表修复语义。
        if line.startswith("index "):
            continue
        if line.startswith("@@"):
            line = re.sub(r"^@@[^@]*@@", "@@ @@", line)
        normalized.append(line.rstrip())
    return "\n".join(normalized).strip()


def patch_features(patch: str) -> PatchFeatures:
    """把 patch 拆成后续指标要用的结构特征。"""
    files: list[str] = []
    changed_lines: list[str] = []
    hunk_labels: list[str] = []
    additions = deletions = 0
    for line in patch.splitlines():
        if line.startswith("diff --git a/") and " b/" in line:
            files.append(line.split(" b/", 1)[1].strip())
        elif line.startswith("@@"):
            match = HUNK_RE.match(line)
            label = match.group(1).strip() if match else ""
            if label:
                # hunk label 通常是函数名或上下文标签，可辅助判断是否落在同一代码区域。
                hunk_labels.append(label)
        elif line.startswith("+") and not line.startswith("+++"):
            additions += 1
            changed_lines.append("+" + normalize_code_line(line[1:]))
        elif line.startswith("-") and not line.startswith("---"):
            deletions += 1
            changed_lines.append("-" + normalize_code_line(line[1:]))
    # token 指标用于捕捉变量名、函数名、常量等代码元素是否接近。
    tokens = TOKEN_RE.findall("\n".join(changed_lines))
    normalized = normalize_patch(patch)
    return PatchFeatures(
        valid=bool(files and changed_lines),
        files=sorted(set(files)),
        additions=additions,
        deletions=deletions,
        changed_lines=changed_lines,
        tokens=tokens,
        hunk_labels=sorted(set(hunk_labels)),
        normalized_patch=normalized,
        patch_id=stable_patch_id(patch),
    )


def jaccard(left: Iterable[str], right: Iterable[str]) -> float:
    """集合相似度：两个集合交集 / 并集。这里主要用于文件集合和 hunk 标签。"""
    a, b = set(left), set(right)
    if not a and not b:
        return 1.0
    return len(a & b) / len(a | b) if a | b else 0.0


def multiset_f1(left: Iterable[str], right: Iterable[str]) -> float:
    """多重集合 F1：同时考虑人工变更覆盖率和 AI 变更精确率。"""
    a, b = Counter(left), Counter(right)
    if not a and not b:
        return 1.0
    overlap = sum((a & b).values())
    precision = overlap / sum(b.values()) if b else 0.0
    recall = overlap / sum(a.values()) if a else 0.0
    return 2 * precision * recall / (precision + recall) if precision + recall else 0.0


def multiset_precision_recall(reference: Iterable[str], candidate: Iterable[str]) -> tuple[float, float]:
    """返回 (AI 变更精确率, 人工变更覆盖率)。

    reference 是人工 patch 的变更行，candidate 是 AI patch 的变更行。
    coverage/recall 低通常意味着漏改；precision 低通常意味着多改或乱改。
    使用 Counter 是为了保留重复变更行的出现次数，而不是简单去重。
    """
    expected, actual = Counter(reference), Counter(candidate)
    if not expected and not actual:
        return 1.0, 1.0
    overlap = sum((expected & actual).values())
    precision = overlap / sum(actual.values()) if actual else 0.0
    recall = overlap / sum(expected.values()) if expected else 0.0
    return precision, recall


def ratio_similarity(left: int, right: int) -> float:
    """规模相似度：用于比较人工 patch 与 AI patch 的增删行数量是否接近。"""
    if left == right == 0:
        return 1.0
    return min(left, right) / max(left, right) if max(left, right) else 0.0


def compare(manual_patch: str, ai_patch: str, result: dict[str, Any]) -> dict[str, Any]:
    manual = patch_features(manual_patch)
    ai = patch_features(ai_patch)
    # tree_match_manual 是上游评测脚本已经在同一 baseline 上算出的结果等价证据。
    tree_match = truthy(result.get("tree_match_manual"))
    patch_id_match = bool(manual.patch_id and manual.patch_id == ai.patch_id)
    exact_normalized = bool(manual.normalized_patch and manual.normalized_patch == ai.normalized_patch)

    # 文件集合一致性回答“改的是不是同一批文件”。
    file_score = jaccard(manual.files, ai.files)

    # changed_line_f1 是严格文本级变更相似度；空行、括号行等格式变化也会影响它。
    line_score = multiset_f1(manual.changed_lines, ai.changed_lines)
    line_precision, line_recall = multiset_precision_recall(manual.changed_lines, ai.changed_lines)

    # token/hunk/diffstat/sequence 指标用于补充文本行完全匹配之外的结构相似度。
    token_score = multiset_f1(manual.tokens, ai.tokens)
    hunk_score = jaccard(manual.hunk_labels, ai.hunk_labels)
    stat_score = (
        ratio_similarity(manual.additions, ai.additions)
        + ratio_similarity(manual.deletions, ai.deletions)
    ) / 2
    sequence_score = SequenceMatcher(
        None, "\n".join(manual.changed_lines), "\n".join(ai.changed_lines), autojunk=False
    ).ratio()

    # structural_score 是机械初筛总分，不代表漏洞修复正确。
    # 权重更偏向文件命中和变更行匹配，其次看顺序、token、hunk 标签和增删规模。
    structural_score = round(
        0.25 * file_score
        + 0.25 * line_score
        + 0.15 * sequence_score
        + 0.15 * token_score
        + 0.10 * hunk_score
        + 0.10 * stat_score,
        4,
    )
    files_exact = manual.files == ai.files
    missing_files = sorted(set(manual.files) - set(ai.files))
    extra_files = sorted(set(ai.files) - set(manual.files))

    # verdict 是分层机械判定：强证据优先，弱相似进入语义审查。
    if not manual.valid or not ai.valid:
        verdict = "insufficient_data"
    elif tree_match is True:
        verdict = "result_tree_match"
    elif patch_id_match:
        verdict = "patch_id_equivalent"
    elif exact_normalized:
        verdict = "normalized_diff_match"
    elif files_exact and structural_score >= 0.85:
        verdict = "likely_equivalent_needs_validation"
    elif missing_files and file_score < 0.5:
        verdict = "divergent"
    elif structural_score >= 0.55:
        verdict = "semantic_review_required"
    else:
        verdict = "divergent"

    return {
        "source_commit": result.get("source_commit", ""),
        "source_repo": result.get("source_repo", ""),
        "target_repo": result.get("target_repo", ""),
        "manual_commit": result.get("manual_commit", ""),
        "case_baseline": result.get("case_baseline", ""),
        "expected_behavior": result.get("expected_behavior", ""),
        "decision": result.get("decision", ""),
        "apply_success": truthy(result.get("apply_success")),
        "tree_match_manual": tree_match,
        "patch_id_match": patch_id_match,
        "exact_normalized_diff": exact_normalized,
        "manual_files": manual.files,
        "ai_files": ai.files,
        "missing_manual_files": missing_files,
        "extra_ai_files": extra_files,
        "files_exact": files_exact,
        "file_set_jaccard": round(file_score, 4),
        "changed_line_f1": round(line_score, 4),
        # manual_change_coverage = recall：人工 patch 的变更被 AI patch 覆盖了多少。
        "manual_change_coverage": round(line_recall, 4),
        # ai_change_precision = precision：AI patch 的变更有多少能对上人工 patch。
        "ai_change_precision": round(line_precision, 4),
        "changed_sequence_similarity": round(sequence_score, 4),
        "token_f1": round(token_score, 4),
        "hunk_label_jaccard": round(hunk_score, 4),
        "diffstat_similarity": round(stat_score, 4),
        "structural_score": structural_score,
        "manual_additions": manual.additions,
        "manual_deletions": manual.deletions,
        "ai_additions": ai.additions,
        "ai_deletions": ai.deletions,
        "verdict": verdict,
        "data_issue": (
            "manual patch missing or malformed"
            if not manual.valid
            else "AI patch missing or malformed"
            if not ai.valid
            else ""
        ),
    }


def write_outputs(output_dir: Path, evaluations: list[dict[str, Any]], review_packets: list[dict[str, Any]]) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "similarity.json").write_text(
        json.dumps(evaluations, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    flat_columns = [
        "source_commit", "source_repo", "target_repo", "manual_commit",
        "case_baseline", "expected_behavior",
        "decision", "apply_success",
        "tree_match_manual", "patch_id_match", "files_exact", "file_set_jaccard",
        "changed_line_f1", "manual_change_coverage", "ai_change_precision",
        "changed_sequence_similarity", "token_f1",
        "hunk_label_jaccard", "diffstat_similarity", "structural_score", "verdict",
    ]
    with (output_dir / "similarity.csv").open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=flat_columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(evaluations)
    with (output_dir / "llm_review.jsonl").open("w", encoding="utf-8") as handle:
        for packet in review_packets:
            handle.write(json.dumps(packet, ensure_ascii=False) + "\n")
    counts = Counter(row["verdict"] for row in evaluations)
    lines = ["# Patch Similarity Summary", "", f"Cases: {len(evaluations)}", ""]
    lines.extend(f"- `{name}`: {count}" for name, count in sorted(counts.items()))
    lines.extend(["", "Similarity is evidence, not proof of correctness."])
    (output_dir / "summary.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Layered similarity evaluation for ai_backport_eval workbooks.",
        epilog=WORKBOOK_CONTRACT,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output-dir", type=Path, default=Path("patch-similarity-output"))
    parser.add_argument("--decision", default="mystique_patch", help="Results decision to evaluate; use 'all' for all")
    parser.add_argument(
        "--expected-behavior",
        default="manual_backport",
        help=(
            "Results expected_behavior to evaluate; use 'all' for all. "
            "Rows without this column/value are kept for backward compatibility."
        ),
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    workbook = openpyxl.load_workbook(args.workbook, read_only=True, data_only=True)
    repo_context = run_info_values(workbook)
    results = reassemble_chunked_rows(sheet_rows(workbook["Results"]), "source_commit")
    conflicts = reassemble_chunked_rows(sheet_rows(workbook["ConflictDetails"]), "source_commit")
    result_by_source = {str(row.get("source_commit")): row for row in results if row.get("source_commit")}
    evaluations: list[dict[str, Any]] = []
    review_packets: list[dict[str, Any]] = []
    for detail in conflicts:
        source = str(detail.get("source_commit") or "")
        result = {
            **repo_context,
            **result_by_source.get(source, {"source_commit": source}),
        }
        expected_behavior = str(result.get("expected_behavior") or "")
        if (
            args.expected_behavior != "all"
            and expected_behavior
            and expected_behavior != args.expected_behavior
        ):
            continue
        if args.decision != "all" and result.get("decision") != args.decision:
            continue
        manual_patch = str(detail.get("manual_patch") or "")
        ai_patch = str(detail.get("mystique_backported_patch") or "")
        evaluation = compare(manual_patch, ai_patch, result)
        evaluations.append(evaluation)
        if evaluation["verdict"] in {"semantic_review_required", "likely_equivalent_needs_validation", "divergent"}:
            review_packets.append({
                "source_commit": source,
                "source_repo": evaluation.get("source_repo", ""),
                "target_repo": evaluation.get("target_repo", ""),
                "mechanical_evidence": evaluation,
                "manual_patch": manual_patch,
                "ai_patch": ai_patch,
            })
    write_outputs(args.output_dir, evaluations, review_packets)
    print(json.dumps({"cases": len(evaluations), "output_dir": str(args.output_dir)}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
